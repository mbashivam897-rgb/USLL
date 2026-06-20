"""Extend USL model forecast horizon FY2027-2030 -> FY2027-2034 (8 years),
preserving all formulas, the waterfall depreciation method, formatting and references."""
import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy

SRC = "/projects/sandbox/USLL/United Spirits Excell.xlsx"
OUT = "/projects/sandbox/USLL/United Spirits Excell - 8 Year Forecast.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=False)

def cstyle(dst, src):
    if src.has_style:
        dst._style = copy(src._style)

def setf(ws, col, row, formula, style_src_cell=None):
    c = ws.cell(row=row, column=column_index_from_string(col))
    c.value = formula
    if style_src_cell is not None:
        cstyle(c, style_src_cell)
    return c

def extend_cols(ws, src_col, new_cols, row_start=1, row_end=None):
    """Append new year columns by translating the last forecast column's formulas
    (relative refs shift; absolute $refs stay)."""
    sidx = column_index_from_string(src_col)
    row_end = row_end or ws.max_row
    for r in range(row_start, row_end + 1):
        src = ws.cell(row=r, column=sidx)
        for ncl in new_cols:
            dst = ws.cell(row=r, column=column_index_from_string(ncl))
            cstyle(dst, src)
            v = src.value
            if isinstance(v, str) and v.startswith("="):
                dst.value = Translator(v, origin=f"{src_col}{r}").translate_formula(f"{ncl}{r}")
            elif v is not None and not isinstance(v, str):
                dst.value = v
            elif isinstance(v, str):
                dst.value = v
    for ncl in new_cols:
        ws.column_dimensions[ncl].width = ws.column_dimensions[src_col].width

# ---------- sanity check on Translator ----------
pl = wb["PL"]
print("SANITY translate PL P6 ->",
      Translator(pl["P6"].value, origin="P6").translate_formula("Q6"))
print("SANITY translate PL P10 ->",
      Translator(pl["P10"].value, origin="P10").translate_formula("Q10"))

# =====================================================================
# 1) SCENARIOS  (years H:K -> append L,M,N,O)
# =====================================================================
sc = wb["Scenarios"]
extend_cols(sc, "K", ["L", "M", "N", "O"], 1, 29)
# Override the scenario driver values for 2031-2034 (fade growth toward ~6% terminal)
scen_vals = {
    12: [0.07, 0.065, 0.06, 0.06],     # Base sales growth
    13: [0.08, 0.07, 0.07, 0.06],      # Best
    14: [0.05, 0.045, 0.04, 0.04],     # Bear
    21: [0.47, 0.475, 0.475, 0.48],    # Base GP margin
    22: [0.515, 0.52, 0.52, 0.52],     # Best
    23: [0.44, 0.44, 0.44, 0.44],      # Bear
}
for row, vals in scen_vals.items():
    for col, val in zip(["L", "M", "N", "O"], vals):
        sc[f"{col}{row}"] = val

# =====================================================================
# 2) ASSUMPTIONS  (years E:H, notes in col I -> relocate notes to N; append I:L)
# =====================================================================
asm = wb["Assumptions"]
# relocate notes column I -> N
for r in range(4, 28):
    src = asm.cell(row=r, column=column_index_from_string("I"))
    if src.value is not None:
        dst = asm.cell(row=r, column=column_index_from_string("N"))
        dst.value = src.value
        cstyle(dst, src)
        src.value = None
asm.column_dimensions["N"].width = asm.column_dimensions["I"].width
# now append year columns I,J,K,L from H
extend_cols(asm, "H", ["I", "J", "K", "L"], 4, 27)

# =====================================================================
# 3) PL, BS, CS, Working Capital Schedule  (years M:P -> append Q,R,S,T)
# =====================================================================
for sheet in ["PL", "BS", "CS", "Working Capital Schedule"]:
    extend_cols(wb[sheet], "P", ["Q", "R", "S", "T"])

# =====================================================================
# 4) WORKING SCHEDULES  (years H:K -> append L,M,N,O; rebuild dep waterfalls)
# =====================================================================
ws = wb["Working Schedules"]
COLS = ["H", "I", "J", "K", "L", "M", "N", "O"]   # 8 forecast years 2027..2034

# capture style templates BEFORE clearing
sty = {
    "sec": ws["B85"], "partic": ws["B86"], "yrhdr": ws["H86"],
    "label": ws["B88"], "open": ws["H88"], "dep": ws["H90"],
    "vint": ws["H91"], "tot": ws["H95"], "close": ws["H96"],
    "ul_lab": ws["M86"], "ul_val": ws["N86"],
}
# generic-extend the UPPER schedules rows 1..82 (lease/debt/finance/CWIP/ROU asset/PPE/intang)
extend_cols(ws, "K", ["L", "M", "N", "O"], 1, 82)

# clear depreciation region rows 85..117 in cols B..O (keep scratch tables in cols P+).
# Preserve only the PPE 'Particulars' header row 86 historical year labels in C:G.
for r in range(85, 118):
    for col in ["B", "C", "D", "E", "F", "G"] + COLS:
        if r == 86 and col in ("C", "D", "E", "F", "G"):
            continue
        ws.cell(row=r, column=column_index_from_string(col)).value = None

# ---- relocate useful-life helpers off the year columns ----
setf(ws, "Q", 86, "Useful life", sty["ul_lab"]); setf(ws, "R", 86, 16, sty["ul_val"])
setf(ws, "Q", 103, "Useful life", sty["ul_lab"]); setf(ws, "R", 103, 7, sty["ul_val"])

# ---- labels ----
labels = {85:("B","Depreciation on PPE Schedule",sty["sec"]),86:("B","Particulars",sty["partic"]),
          88:("B","Opening",sty["label"]),89:("B","Capex",sty["label"]),90:("B","Depreciation",sty["label"]),
          99:("B","Depreciation",sty["label"]),100:("B","Closing",sty["label"]),
          102:("B","Depreciation on ROU Schedule",sty["sec"]),103:("B","Particulars",sty["partic"]),
          105:("B","Opening",sty["label"]),106:("B","Capex",sty["label"]),107:("B","Depreciation",sty["label"]),
          116:("B","Depreciation",sty["label"]),117:("B","Closing",sty["label"])}
for i in range(8):
    labels[91+i] = ("B", f"Dep year {i+1} capex", sty["label"])
    labels[108+i] = ("B", f"Dep year {i+1} capex", sty["label"])
for row,(col,txt,st) in labels.items():
    setf(ws, col, row, txt, st)

# ---- year header rows 86 (PPE) and 103 (ROU) ----
setf(ws, "H", 86, "=G86+1", sty["yrhdr"])
for k in range(1, 8):
    setf(ws, COLS[k], 86, f"={COLS[k-1]}86+1", sty["yrhdr"])
# ROU header row 103: rebuild full year chain C..O so it is self-contained
setf(ws, "C", 103, 2022, sty["yrhdr"])
for col_prev, col in zip(["C","D","E","F","G","H","I","J","K","L","M","N"],
                         ["D","E","F","G","H","I","J","K","L","M","N","O"]):
    setf(ws, col, 103, f"={col_prev}103+1", sty["yrhdr"])

# ---- PPE depreciation waterfall (UL=$R$86) ----
for k, c in enumerate(COLS):
    prev = COLS[k-1] if k > 0 else "G"
    setf(ws, c, 88, f"={prev}67", sty["open"])               # Opening = prior net block
    setf(ws, c, 89, f"={c}63", sty["open"])                  # Capex (gross PPE capex)
    setf(ws, c, 90, (f"=-{c}88/$R$86" if k == 0 else f"={prev}90"), sty["dep"])
for s in range(8):                                            # vintage layers yr1..yr8
    row = 91 + s
    for k, c in enumerate(COLS):
        if k < s:   continue
        setf(ws, c, row, (f"=-{c}89/$R$86" if k == s else f"={COLS[k-1]}{row}"), sty["vint"])
for c in COLS:
    setf(ws, c, 99,  f"=SUM({c}90:{c}98)", sty["tot"])        # Total depreciation
    setf(ws, c, 100, f"=SUM({c}88:{c}89)+{c}99", sty["close"])# Closing

# ---- ROU depreciation waterfall (UL=$R$103) ----
for k, c in enumerate(COLS):
    prev = COLS[k-1] if k > 0 else "G"
    setf(ws, c, 105, (f"=G55" if k == 0 else f"={prev}117"), sty["open"])
    setf(ws, c, 106, f"={c}52", sty["open"])
    setf(ws, c, 107, (f"=-{c}105/$R$103" if k == 0 else f"={prev}107"), sty["dep"])
for s in range(8):
    row = 108 + s
    for k, c in enumerate(COLS):
        if k < s:   continue
        setf(ws, c, row, (f"=-{c}106/$R$103" if k == s else f"={COLS[k-1]}{row}"), sty["vint"])
for c in COLS:
    setf(ws, c, 116, f"=SUM({c}107:{c}115)", sty["tot"])
    setf(ws, c, 117, f"=SUM({c}105:{c}106)+{c}116", sty["close"])

# ---- OVERRIDE cross-refs that pointed at old total rows (95/108 -> 99/116) ----
sty53, sty66, sty70 = ws["H53"], ws["H66"], ws["H70"]
for k, c in enumerate(COLS):
    prev = COLS[k-1] if k > 0 else "G"
    setf(ws, c, 53, f"={c}116", sty53)                        # ROU dep in asset schedule
    setf(ws, c, 66, f"={prev}66+{c}99", sty66)                # PPE accumulated depreciation
    setf(ws, c, 70, f"=-{c}82+{c}99+{c}116", sty70)           # Total D&A
ws.column_dimensions["L"].width = ws.column_dimensions["K"].width
for cc in ["M","N","O"]:
    ws.column_dimensions[cc].width = ws.column_dimensions["K"].width

# =====================================================================
# 5) DCF  (years C:F -> append G,H,I,J; fix terminal & discounting)
# =====================================================================
dcf = wb["DCF"]
extend_cols(dcf, "F", ["G", "H", "I", "J"], 1, 15)
for col, per in zip(["G", "H", "I", "J"], [5, 6, 7, 8]):
    dcf[f"{col}13"] = per                                     # discount periods
dcf["C19"] = "=J12"                                           # terminal-year FCF (last explicit yr)
dcf["C21"] = "=C20/(1+N9)^J13"                                # PV of terminal value (period 8)
dcf["C22"] = "=SUM(C15:J15)"                                  # sum of PV of explicit FCF

# force Excel to recalculate on open (openpyxl writes formulas without cached values)
wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("SAVED", OUT)
