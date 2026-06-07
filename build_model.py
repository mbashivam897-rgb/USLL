"""
United Spirits Ltd (USL) - Debt Schedule & Cash Schedule
Consolidated, Rs crore. FY2022A-FY2025A actuals from FY2023 & FY2025 Annual Reports.
FY2026E-FY2030E forecast. Interest is computed via live formulas.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---------- Styling palette ----------
NAVY   = "1F3864"
BLUE   = "2E5496"
LBLUE  = "D9E1F2"
GREY   = "F2F2F2"
GOLD   = "BF9000"
GREEN  = "548235"
INPUTC = "0000CC"   # blue font for inputs
ORANGE = "FFF2CC"
WHITE  = "FFFFFF"
DARKR  = "C00000"

thin = Side(style="thin", color="BFBFBF")
medium = Side(style="medium", color="808080")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
top_border = Border(top=Side(style="thin", color="404040"))
dbl_top = Border(top=Side(style="double", color="404040"))

def style_title(ws, cell, text, size=16):
    ws[cell] = text
    ws[cell].font = Font(name="Calibri", size=size, bold=True, color=WHITE)
    ws[cell].fill = PatternFill("solid", fgColor=NAVY)
    ws[cell].alignment = Alignment(vertical="center", horizontal="left", indent=1)

def hdr_fill(c, color=BLUE, fontcolor=WHITE, bold=True, size=10, align="center"):
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=bold, color=fontcolor, size=size)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = border_all

YEARS = ["FY22A", "FY23A", "FY24A", "FY25A", "FY26E", "FY27E", "FY28E", "FY29E", "FY30E"]
# columns C..K
COLS = [get_column_letter(3 + i) for i in range(len(YEARS))]   # C,D,...,K
FCOLS = COLS[4:]   # forecast cols G..K
HCOLS = COLS[:4]   # hist cols C..F

NUMFMT = '#,##0.0;(#,##0.0)'
NUMFMT0 = '#,##0;(#,##0)'
PCT = '0.0%'
PCT2 = '0.00%'

def write_year_header(ws, row, label_widthcols=2, start_text="Rs crore"):
    ws.cell(row=row, column=1, value=start_text)
    c = ws.cell(row=row, column=1); c.font = Font(italic=True, size=8, color="808080")
    for i, (col, yr) in enumerate(zip(COLS, YEARS)):
        cell = ws[f"{col}{row}"]
        cell.value = yr
        color = GREEN if "E" not in yr else GOLD
        hdr_fill(cell, color=(BLUE if yr.endswith("A") else GOLD))
    # mark actual vs estimate row above
    return row

def set_widths(ws, label_w=46, note_w=2, data_w=11):
    ws.column_dimensions["A"].width = label_w
    ws.column_dimensions["B"].width = note_w
    for col in COLS:
        ws.column_dimensions[col].width = data_w

def label(ws, row, text, bold=False, indent=1, italic=False, color="000000", size=10):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=bold, italic=italic, color=color, size=size)
    c.alignment = Alignment(indent=indent, vertical="center")
    return c

def section(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHITE, size=10)
    c.fill = PatternFill("solid", fgColor=BLUE)
    for col in COLS:
        ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=BLUE)
    return c

def put(ws, col, row, val, fmt=NUMFMT, inp=False, bold=False, color=None, indent=None):
    cell = ws[f"{col}{row}"]
    cell.value = val
    cell.number_format = fmt
    fc = INPUTC if inp else (color if color else "000000")
    cell.font = Font(bold=bold, color=fc)
    cell.alignment = Alignment(horizontal="right")
    return cell

def putrow(ws, row, vals, fmt=NUMFMT, inp=False, bold=False, color=None, cols=COLS):
    for col, v in zip(cols, vals):
        if v is not None:
            put(ws, col, row, v, fmt=fmt, inp=inp, bold=bold, color=color)

# =====================================================================
# SHEET 1: COVER
# =====================================================================
cover = wb.active
cover.title = "Cover"
cover.sheet_view.showGridLines = False
set_widths(cover, label_w=24, data_w=14)
for r in range(1, 60):
    pass
cover.merge_cells("A1:K2")
style_title(cover, "A1", "  UNITED SPIRITS LIMITED (USL)  |  Debt Schedule & Cash Schedule", size=18)
cover.merge_cells("A3:K3")
cover["A3"] = "Consolidated  |  Interest computation model  |  All figures in Rs crore unless stated"
cover["A3"].font = Font(italic=True, size=11, color=NAVY)

rows_cover = [
    ("", ""),
    ("Prepared by", "Equity Research – Financial Modelling"),
    ("Company", "United Spirits Limited (a Diageo Group company), NSE/BSE: UNITDSPR / 532432"),
    ("Reporting currency", "Indian Rupee (Rs), figures in Rs crore (1 crore = 10 million)"),
    ("Historical period", "FY2022 to FY2025 (year ended 31 March) - Actuals"),
    ("Forecast period", "FY2026 to FY2030 (year ended 31 March) - Estimates"),
    ("Primary sources", "USL Integrated Annual Report FY2022-23 (figures in Rs million) and FY2024-25 (figures in Rs crore)"),
    ("", ""),
    ("KEY INSIGHT", "USL is structurally NET-CASH / debt-free. Its only contractual borrowings are negligible;"),
    ("", "the material interest-bearing item is LEASE LIABILITIES under Ind AS 116."),
    ("", "Per FY23 AR: \"As the Company is debt-free, exposure to interest rate risk is negligible.\""),
    ("", ""),
    ("Workbook structure", ""),
    ("  1. Cover", "This sheet"),
    ("  2. Assumptions", "All forecast drivers (blue font = input) with justifications"),
    ("  3. Historicals", "Reconstructed consolidated financials FY22A-FY25A from the annual reports"),
    ("  4. Debt Schedule", "Lease liabilities + borrowings roll-forward; computes INTEREST EXPENSE / finance cost"),
    ("  5. Cash Schedule", "Free-cash-flow build, treasury (cash+deposits+MF) roll; computes INTEREST INCOME"),
    ("  6. Output Summary", "Total debt, net debt, net interest, leverage & coverage ratios"),
    ("", ""),
    ("Convention", "Lease interest = rate x average lease balance. Interest income = yield x opening treasury"),
    ("", "(both non-circular). Blue figures are hard-coded inputs; black figures are formulas."),
]
r = 5
for a, b in rows_cover:
    cover.cell(row=r, column=1, value=a).font = Font(bold=(a not in ("", "  1. Cover","  2. Assumptions","  3. Historicals","  4. Debt Schedule","  5. Cash Schedule","  6. Output Summary")), color=NAVY if a.isupper() and a else "000000")
    cover.cell(row=r, column=1).alignment = Alignment(horizontal="left")
    cover.merge_cells(f"B{r}:K{r}")
    cell = cover.cell(row=r, column=2, value=b)
    cell.alignment = Alignment(horizontal="left", wrap_text=False)
    if a == "KEY INSIGHT":
        cover.cell(row=r, column=1).fill = PatternFill("solid", fgColor=ORANGE)
        cell.font = Font(bold=True, color=DARKR)
    r += 1

# =====================================================================
# SHEET 2: ASSUMPTIONS
# =====================================================================
aw = wb.create_sheet("Assumptions")
aw.sheet_view.showGridLines = False
set_widths(aw, label_w=46, data_w=11)
aw.merge_cells("A1:K1")
style_title(aw, "A1", "  ASSUMPTIONS  (blue = input)   -   Rs crore unless stated")
HR = 3
write_year_header(aw, HR)
aw.column_dimensions["L"].width = 70
aw["L"+str(HR)] = "Justification / basis"
hdr_fill(aw["L"+str(HR)], color=GOLD)

# Row map for assumptions (we will reference these cells from schedules)
A = {}
r = HR + 1
def arow(key, text, vals, fmt=NUMFMT, inp=True, note="", bold=False, cols=COLS):
    global r
    label(aw, r, text, bold=bold)
    putrow(aw, r, vals, fmt=fmt, inp=inp, cols=cols)
    aw.cell(row=r, column=12, value=note).alignment = Alignment(wrap_text=True, vertical="center")
    aw.cell(row=r, column=12).font = Font(size=8, color="595959")
    A[key] = r
    r += 1
    return r-1

section(aw, r, "OPERATING DRIVERS"); r += 1
# Net sales actuals then growth inputs
arow("netsales", "Net sales (ex-excise) - actual",
     [9712.4, 10611.6, 11321, 12069, None, None, None, None, None],
     note="From P&L: Revenue from operations less Excise duty. FY22-23 converted from Rs mn (AR FY23); FY24-25 from AR FY25.")
arow("growth", "  Net sales growth %",
     ["=C{0}/C{0}".format(A["netsales"]) and None,  # placeholder
      "=D{0}/C{0}-1".format(A["netsales"]),
      "=E{0}/D{0}-1".format(A["netsales"]),
      "=F{0}/E{0}-1".format(A["netsales"]),
      0.095, 0.10, 0.10, 0.09, 0.09],
     fmt=PCT,
     note="FY26-30 inputs. USL reaffirmed double-digit topline guidance; premiumization (P&A ~90% of mix). Industry premium-spirits value CAGR ~12%, IMFL value ~13%; kept conservative for state excise/regulatory risk.")
arow("ebitdam", "  EBITDA margin (% of net sales)",
     [0.166, 0.133, 0.177, 0.185, 0.190, 0.195, 0.200, 0.205, 0.210],
     fmt=PCT,
     note="Hist derived (FY23 dip = input-cost inflation). FY26-30 expansion on premiumization & operating leverage (H1FY26 reported EBITDA +31%).")
arow("dep", "Depreciation & amortisation",
     [303.8, 282.5, 275, 283, 290, 300, 315, 330, 345], inp=False,
     note="Hist actual. Forecast grows with capex & rising ROU asset base (Ind AS 116).")
for c in HCOLS:
    aw[f"{c}{A['dep']}"].font = Font(color="000000")
for c in FCOLS:
    aw[f"{c}{A['dep']}"].font = Font(color=INPUTC)
arow("tax", "Effective tax rate %",
     [0.254, 0.119, 0.242, 0.259, 0.255, 0.255, 0.255, 0.255, 0.255], fmt=PCT,
     note="Hist actual (FY23 low due to prior-year credits). ~25.5% normalised (India corporate rate).")
arow("payout", "Dividend payout (% of PAT)",
     [0.0, 0.0, 0.202, 0.224, 0.30, 0.30, 0.30, 0.30, 0.30], fmt=PCT,
     note="USL resumed dividends FY24. Rising payout assumed as company is cash-generative & net-cash.")
arow("wcpct", "Working-capital outflow (% of net-sales increase)",
     [None, None, None, None, 0.05, 0.05, 0.05, 0.05, 0.05], fmt=PCT,
     note="Incremental net working-capital investment as sales grow (inventory/receivables drag), funded internally.")
arow("capex", "Capex (PP&E + intangibles)",
     [134.0, 136.6, 150, 160, 175, 200, 225, 250, 250],
     note="Hist actual/estimate. Forecast: capacity, premiumization & supply-restructuring; asset-light model keeps capex modest.")
for c in HCOLS:
    aw[f"{c}{A['capex']}"].font = Font(color="000000")
for c in FCOLS:
    aw[f"{c}{A['capex']}"].font = Font(color=INPUTC)

r += 1
section(aw, r, "DEBT / LEASE DRIVERS"); r += 1
arow("leaserate", "Lease incremental borrowing rate %",
     [0.085, 0.085, 0.085, 0.085, 0.085, 0.085, 0.085, 0.085, 0.085], fmt=PCT,
     note="Ind AS 116 discount rate. Implied FY25 lease interest 40 / avg lease ~360 = ~8-9%. Held at 8.5%.")
arow("leaseadd", "New lease additions (ROU)",
     [229.2, 85.8, 184, 394, 200, 190, 190, 200, 210],
     note="Hist actual (FY25 spike 394 = plant & equipment leases). Forecast normalised to recurring warehouse/plant leases.")
for c in HCOLS:
    aw[f"{c}{A['leaseadd']}"].font = Font(color="000000")
for c in FCOLS:
    aw[f"{c}{A['leaseadd']}"].font = Font(color=INPUTC)
arow("leaserepay", "Lease principal repayment",
     [100.2, 124.0, 126, 137, 150, 160, 170, 180, 190],
     note="Hist actual (financing CF). Forecast grows with the larger lease book.")
for c in HCOLS:
    aw[f"{c}{A['leaserepay']}"].font = Font(color="000000")
for c in FCOLS:
    aw[f"{c}{A['leaserepay']}"].font = Font(color=INPUTC)
arow("leaseterm", "Lease terminations / adjustments",
     [0, 0, 0, 17, 0, 0, 0, 0, 0],
     note="FY25 had Rs 17 cr lease terminations (ROU note). None assumed forward.")
for c in COLS:
    aw[f"{c}{A['leaseterm']}"].font = Font(color="000000" if c in HCOLS else INPUTC)
arow("borrowopen22", "Borrowings - opening (FY22 only)",
     [497.0, None, None, None, None, None, None, None, None],
     note="FY22 opening short-term/working-capital borrowings (repaid during FY22). Backed out from FY22 repayment & closing.")
arow("borrowdraw", "Borrowings drawdown",
     [0, 0, 25, 28, 0, 0, 0, 0, 0],
     note="Hist actual working-capital loan draws. Nil forward: USL follows a debt-free policy with large undrawn facilities (~Rs 1,300 cr).")
for c in COLS:
    aw[f"{c}{A['borrowdraw']}"].font = Font(color="000000" if c in HCOLS else INPUTC)
arow("borrowrepay", "Borrowings repayment",
     [537.7, 0.4, 0, 53, 0, 0, 0, 0, 0],
     note="Hist actual. (FY22 repaid Rs 537.7 cr working-capital loans; FY25 repaid Rs 53 cr.)")
for c in COLS:
    aw[f"{c}{A['borrowrepay']}"].font = Font(color="000000" if c in HCOLS else INPUTC)
arow("borrowrate", "Interest rate on borrowings %",
     [0.0, 0.0, 0.0, 0.0, 0.085, 0.085, 0.085, 0.085, 0.085], fmt=PCT,
     note="Applied to average borrowings (nil forward, so no P&L impact).")
arow("otherint", "Other interest expense (non-lease)",
     [76.1, 87.9, 55, 49, 47, 45, 43, 42, 40],
     note="Interest 'others' in finance-cost note (deferred sales-tax, MSME, unwinding of discount, bank charges). Held ~flat, easing.")
for c in HCOLS:
    aw[f"{c}{A['otherint']}"].font = Font(color="000000")
for c in FCOLS:
    aw[f"{c}{A['otherint']}"].font = Font(color=INPUTC)

r += 1
section(aw, r, "CASH / TREASURY DRIVERS"); r += 1
arow("yield", "Treasury yield (interest income) %",
     [None, None, None, None, 0.0675, 0.0675, 0.0675, 0.0675, 0.0675], fmt=PCT,
     note="Earned on cash + bank deposits + debt mutual funds. ~6.75% reflects Indian short-term debt/MF yields.")
arow("nonopinc", "Non-interest other income (operating)",
     [None, None, None, None, 60, 65, 70, 75, 80],
     note="Recurring portion of 'Other income' excluding treasury interest (scrap sale, write-backs, misc.).")

r += 1
# legend
aw.cell(row=r, column=1, value="Legend:").font = Font(bold=True)
aw.cell(row=r+1, column=1, value="Blue figures = hard-coded inputs/assumptions").font = Font(color=INPUTC, size=9)
aw.cell(row=r+2, column=1, value="Black figures = actuals or formula-driven").font = Font(color="000000", size=9)

ASSUMP = "Assumptions"

# =====================================================================
# SHEET 3: HISTORICALS
# =====================================================================
hw = wb.create_sheet("Historicals")
hw.sheet_view.showGridLines = False
set_widths(hw, label_w=46, data_w=11)
hw.merge_cells("A1:K1")
style_title(hw, "A1", "  HISTORICAL CONSOLIDATED FINANCIALS (as reported)   -   Rs crore")
hw.merge_cells("A2:K2")
hw["A2"] = "Source: USL Annual Reports FY2022-23 (Rs mn, /10) and FY2024-25 (Rs cr). Only actual columns populated."
hw["A2"].font = Font(italic=True, size=9, color="595959")
HR = 4
write_year_header(hw, HR)
H = {}
r = HR + 1
def hrow(key, text, vals, fmt=NUMFMT, bold=False, indent=1, cols=COLS, formula_cells=None):
    global r
    label(hw, r, text, bold=bold, indent=indent)
    putrow(hw, r, vals, fmt=fmt, bold=bold, cols=cols)
    H[key] = r
    r += 1
    return r-1

section(hw, r, "PROFIT & LOSS"); r += 1
hrow("rev", "Revenue from operations (gross, incl excise)", [31061.8, 27815.4, 26018, 27276, None,None,None,None,None])
hrow("excise", "Less: Excise duty", [21349.4, 17203.8, 14697, 15207, None,None,None,None,None])
hrow("ns", "Net sales (ex-excise)", ["=C{0}-C{1}".format(H["rev"],H["excise"]),"=D{0}-D{1}".format(H["rev"],H["excise"]),"=E{0}-E{1}".format(H["rev"],H["excise"]),"=F{0}-F{1}".format(H["rev"],H["excise"]),None,None,None,None,None], bold=True)
hrow("oi", "Other income", [35.5, 73.1, 225, 336, None,None,None,None,None])
hrow("ebitda", "EBITDA (derived, pre-exceptional)", [1608.1, 1415.5, 2000, 2236, None,None,None,None,None], bold=True)
hw.cell(row=H["ebitda"], column=1).font = Font(bold=True)
hrow("ebm", "  EBITDA margin % (of net sales)",
     ["=C{0}/C{1}".format(H["ebitda"],H["ns"]),"=D{0}/D{1}".format(H["ebitda"],H["ns"]),"=E{0}/E{1}".format(H["ebitda"],H["ns"]),"=F{0}/F{1}".format(H["ebitda"],H["ns"]),None,None,None,None,None], fmt=PCT)
hrow("dep", "Depreciation & amortisation", [303.8, 282.5, 275, 283, None,None,None,None,None])
hrow("fc", "Finance costs (total)", [88.0, 103.9, 76, 89, None,None,None,None,None])
hrow("fc_lease", "  of which: lease interest", [11.9, 16.0, 21, 40, None,None,None,None,None], indent=2)
hrow("fc_oth", "  of which: other interest", [76.1, 87.9, 55, 49, None,None,None,None,None], indent=2)
hrow("exc", "Exceptional items (net)", [-165.2, 176.4, -17, -65, None,None,None,None,None])
hrow("pbt", "Profit before tax", [1086.6, 1278.6, 1857, 2135, None,None,None,None,None], bold=True)
hrow("tax", "Total tax expense", [276.0, 152.8, 449, 553, None,None,None,None,None])
hrow("pat", "Profit after tax (PAT)", [810.6, 1125.8, 1408, 1582, None,None,None,None,None], bold=True)

r += 1
section(hw, r, "DEBT, LEASES & LIQUIDITY (Balance Sheet)"); r += 1
hrow("b_nc", "Borrowings - non-current", [0.9, 0.3, 0, 0, None,None,None,None,None])
hrow("b_c", "Borrowings - current", [340.8, 0.8, 25, 0, None,None,None,None,None])
hrow("b_tot", "Total borrowings",
     ["=C{0}+C{1}".format(H["b_nc"],H["b_c"]),"=D{0}+D{1}".format(H["b_nc"],H["b_c"]),"=E{0}+E{1}".format(H["b_nc"],H["b_c"]),"=F{0}+F{1}".format(H["b_nc"],H["b_c"]),None,None,None,None,None], bold=True)
hrow("l_nc", "Lease liabilities - non-current", [134.1, 80.0, 137, 334, None,None,None,None,None])
hrow("l_c", "Lease liabilities - current", [129.6, 102.2, 103, 146, None,None,None,None,None])
hrow("l_tot", "Total lease liabilities",
     ["=C{0}+C{1}".format(H["l_nc"],H["l_c"]),"=D{0}+D{1}".format(H["l_nc"],H["l_c"]),"=E{0}+E{1}".format(H["l_nc"],H["l_c"]),"=F{0}+F{1}".format(H["l_nc"],H["l_c"]),None,None,None,None,None], bold=True)
hrow("debt_tot", "Total debt (borrowings + leases)",
     ["=C{0}+C{1}".format(H["b_tot"],H["l_tot"]),"=D{0}+D{1}".format(H["b_tot"],H["l_tot"]),"=E{0}+E{1}".format(H["b_tot"],H["l_tot"]),"=F{0}+F{1}".format(H["b_tot"],H["l_tot"]),None,None,None,None,None], bold=True)
hw.cell(row=H["debt_tot"], column=1).font = Font(bold=True, color=DARKR)
hrow("cash", "Cash & cash equivalents", [54.5, 115.1, 1052, 1328, None,None,None,None,None])
hrow("bank", "Bank balances other than cash (deposits)", [5.8, 768.2, 217, 702, None,None,None,None,None])
hrow("mf", "Current investments (debt mutual funds)", [222.1, 255.8, 599, 873, None,None,None,None,None])
hrow("treas", "Treasury (cash + deposits + MF)",
     ["=SUM(C{0}:C{1})".format(H["cash"],H["mf"]),"=SUM(D{0}:D{1})".format(H["cash"],H["mf"]),"=SUM(E{0}:E{1})".format(H["cash"],H["mf"]),"=SUM(F{0}:F{1})".format(H["cash"],H["mf"]),None,None,None,None,None], bold=True)
hw.cell(row=H["treas"], column=1).font = Font(bold=True, color=GREEN)
hrow("netdebt", "Net debt / (net cash)",
     ["=C{0}-C{1}".format(H["debt_tot"],H["treas"]),"=D{0}-D{1}".format(H["debt_tot"],H["treas"]),"=E{0}-E{1}".format(H["debt_tot"],H["treas"]),"=F{0}-F{1}".format(H["debt_tot"],H["treas"]),None,None,None,None,None], bold=True)
hrow("equity", "Total equity", [4874.8, 5999.5, 7121, 8104, None,None,None,None,None])

HIST = "Historicals"

# =====================================================================
# SHEET 4: DEBT SCHEDULE
# =====================================================================
dw = wb.create_sheet("Debt Schedule")
dw.sheet_view.showGridLines = False
set_widths(dw, label_w=46, data_w=11)
dw.merge_cells("A1:K1")
style_title(dw, "A1", "  DEBT SCHEDULE  -  Lease liabilities + Borrowings  ->  INTEREST EXPENSE   (Rs crore)")
HR = 3
write_year_header(dw, HR)
D = {}
r = HR + 1

def drow(key, text, fmt=NUMFMT, bold=False, indent=1, color=None):
    global r
    label(dw, r, text, bold=bold, indent=indent, color=(color or "000000"))
    D[key] = r
    r += 1
    return D[key]

# ---- Lease liabilities roll ----
section(dw, r, "A. LEASE LIABILITIES (Ind AS 116)"); r += 1
drow("l_open", "Opening lease liability")
drow("l_add", "(+) New lease additions (ROU)")
drow("l_term", "(-) Terminations / adjustments")
drow("l_repay", "(-) Principal repayment")
drow("l_close", "Closing lease liability", bold=True)
drow("l_int", "Interest on leases (rate x avg balance)", bold=True, color=DARKR)
drow("l_rate_chk", "  implied interest rate %", fmt=PCT, indent=2)

# Lease opening: FY22 opening = FY22 close - add + repay + term (back-solve) OR use prior-year hist.
# We anchor closing to Historicals actuals for hist columns, and roll forward for forecast.
# Opening FY22 = Historicals FY22 close - additions + repay + term
lo, ladd, lterm, lrepay, lclose, lint, lchk = (D["l_open"],D["l_add"],D["l_term"],D["l_repay"],D["l_close"],D["l_int"],D["l_rate_chk"])
for i, col in enumerate(COLS):
    add = f"{ASSUMP}!{col}{A['leaseadd']}"
    term = f"{ASSUMP}!{col}{A['leaseterm']}"
    repay = f"{ASSUMP}!{col}{A['leaserepay']}"
    rate = f"{ASSUMP}!{col}{A['leaserate']}"
    # opening
    if i == 0:
        # back-solve FY22 opening from hist close
        put(dw, col, lo, f"='{HIST}'!C{H['l_tot']}-{add}+{repay}+{term}")
    else:
        prev = COLS[i-1]
        put(dw, col, lo, f"={prev}{lclose}")
    put(dw, col, ladd, f"={add}")
    put(dw, col, lterm, f"=-{term}")
    put(dw, col, lrepay, f"=-{repay}")
    # closing: for hist columns tie to reported; for forecast = roll
    if col in HCOLS:
        put(dw, col, lclose, f"='{HIST}'!{col}{H['l_tot']}", bold=True)
    else:
        put(dw, col, lclose, f"={col}{lo}+{col}{ladd}+{col}{lterm}+{col}{lrepay}", bold=True)
    # interest = rate * average(open, close); hist ties to reported lease interest
    if col in HCOLS:
        put(dw, col, lint, f"='{HIST}'!{col}{H['fc_lease']}", bold=True, color=DARKR)
    else:
        put(dw, col, lint, f"={rate}*AVERAGE({col}{lo},{col}{lclose})", bold=True, color=DARKR)
    put(dw, col, lchk, f"=IF(AVERAGE({col}{lo},{col}{lclose})=0,0,{col}{lint}/AVERAGE({col}{lo},{col}{lclose}))", fmt=PCT, indent=2)

r += 1
# ---- Borrowings roll ----
section(dw, r, "B. BORROWINGS (working-capital / short-term)"); r += 1
drow("b_open", "Opening borrowings")
drow("b_draw", "(+) Drawdowns")
drow("b_repay", "(-) Repayments")
drow("b_close", "Closing borrowings", bold=True)
drow("b_int", "Interest on borrowings (rate x avg)", bold=True, color=DARKR)
bo,bd,brp,bc,bi = D["b_open"],D["b_draw"],D["b_repay"],D["b_close"],D["b_int"]
for i, col in enumerate(COLS):
    draw = f"{ASSUMP}!{col}{A['borrowdraw']}"
    rep = f"{ASSUMP}!{col}{A['borrowrepay']}"
    rate = f"{ASSUMP}!{col}{A['borrowrate']}"
    if i == 0:
        put(dw, col, bo, f"={ASSUMP}!C{A['borrowopen22']}")
    else:
        put(dw, col, bo, f"={COLS[i-1]}{bc}")
    put(dw, col, bd, f"={draw}")
    put(dw, col, brp, f"=-{rep}")
    if col in HCOLS:
        put(dw, col, bc, f"='{HIST}'!{col}{H['b_tot']}", bold=True)
    else:
        put(dw, col, bc, f"={col}{bo}+{col}{bd}+{col}{brp}", bold=True)
    put(dw, col, bi, f"={rate}*AVERAGE({col}{bo},{col}{bc})", bold=True, color=DARKR)

r += 1
# ---- Total finance cost ----
section(dw, r, "C. TOTAL FINANCE COST (P&L interest expense)"); r += 1
drow("tot_lease_int", "Lease interest")
drow("tot_bor_int", "Borrowing interest")
drow("tot_oth_int", "Other interest (non-lease, non-borrowing)")
drow("tot_fc", "Total finance cost", bold=True, color=DARKR)
drow("tot_debt", "Memo: Total debt (closing)", bold=True)
drow("fc_actual", "Memo: Reported finance cost (actual)", indent=2)
drow("fc_var", "Memo: Variance (model - reported)", indent=2)
drow("modelrate", "Memo: model lease interest @ rate x avg", indent=2)
for col in COLS:
    put(dw, col, D["tot_lease_int"], f"={col}{lint}")
    put(dw, col, D["tot_bor_int"], f"={col}{bi}")
    put(dw, col, D["tot_oth_int"], f"={ASSUMP}!{col}{A['otherint']}")
    put(dw, col, D["tot_fc"], f"=SUM({col}{D['tot_lease_int']}:{col}{D['tot_oth_int']})", bold=True, color=DARKR)
    put(dw, col, D["tot_debt"], f"={col}{lclose}+{col}{bc}", bold=True)
# reported actual finance cost for comparison (hist only)
putrow(dw, D["fc_actual"], [88.0,103.9,76,89,None,None,None,None,None], color="808080")
for i, col in enumerate(COLS):
    # variance model vs reported (hist only)
    if col in HCOLS:
        put(dw, col, D["fc_var"], f"={col}{D['tot_fc']}-{col}{D['fc_actual']}", color="808080")
    # memo: pure rate x avg lease interest (methodology shown for all years)
    rate = f"{ASSUMP}!{col}{A['leaserate']}"
    put(dw, col, D["modelrate"], f"={rate}*AVERAGE({col}{lo},{col}{lclose})", color="808080")

DEBT = "Debt Schedule"

# =====================================================================
# SHEET 5: CASH SCHEDULE
# =====================================================================
cw = wb.create_sheet("Cash Schedule")
cw.sheet_view.showGridLines = False
set_widths(cw, label_w=46, data_w=11)
cw.merge_cells("A1:K1")
style_title(cw, "A1", "  CASH SCHEDULE  -  FCF build & treasury roll  ->  INTEREST INCOME   (Rs crore)")
HR = 3
write_year_header(cw, HR)
C_ = {}
r = HR + 1
def crow(key, text, fmt=NUMFMT, bold=False, indent=1, color=None):
    global r
    label(cw, r, text, bold=bold, indent=indent, color=(color or "000000"))
    C_[key] = r
    r += 1
    return C_[key]

section(cw, r, "A. P&L BRIDGE (drives cash)"); r += 1
crow("ns", "Net sales (ex-excise)")
crow("gr", "  growth %", fmt=PCT, indent=2)
crow("ebitda", "EBITDA", bold=True)
crow("dep", "(-) Depreciation & amortisation")
crow("ebit", "EBIT", bold=True)
crow("intinc", "(+) Interest income (treasury)", color=GREEN)
crow("fc", "(-) Finance cost (from Debt Schedule)", color=DARKR)
crow("nonop", "(+) Other operating income")
crow("pbt", "Profit before tax", bold=True)
crow("tax", "(-) Tax")
crow("pat", "Profit after tax (PAT)", bold=True)

for i, col in enumerate(COLS):
    if col in HCOLS:
        put(cw, col, C_["ns"], f"='{HIST}'!{col}{H['ns']}")
    else:
        prev = COLS[i-1]
        put(cw, col, C_["ns"], f"={prev}{C_['ns']}*(1+{ASSUMP}!{col}{A['growth']})")
    if i == 0:
        put(cw, col, C_["gr"], None, fmt=PCT)
    else:
        put(cw, col, C_["gr"], f"={col}{C_['ns']}/{COLS[i-1]}{C_['ns']}-1", fmt=PCT, indent=2)
    # EBITDA
    if col in HCOLS:
        put(cw, col, C_["ebitda"], f"='{HIST}'!{col}{H['ebitda']}", bold=True)
        put(cw, col, C_["dep"], f"=-'{HIST}'!{col}{H['dep']}")
    else:
        put(cw, col, C_["ebitda"], f"={col}{C_['ns']}*{ASSUMP}!{col}{A['ebitdam']}", bold=True)
        put(cw, col, C_["dep"], f"=-{ASSUMP}!{col}{A['dep']}")
    put(cw, col, C_["ebit"], f"={col}{C_['ebitda']}+{col}{C_['dep']}", bold=True)
    # interest income (forecast formula; hist actual approx blank/treasury*yield not needed)
    if col in HCOLS:
        put(cw, col, C_["intinc"], None, color=GREEN)
        put(cw, col, C_["fc"], f"=-'{HIST}'!{col}{H['fc']}", color=DARKR)
        put(cw, col, C_["nonop"], None)
        put(cw, col, C_["pbt"], f"='{HIST}'!{col}{H['pbt']}", bold=True)
        put(cw, col, C_["tax"], f"=-'{HIST}'!{col}{H['tax']}")
        put(cw, col, C_["pat"], f"='{HIST}'!{col}{H['pat']}", bold=True)
    else:
        # interest income references treasury opening on this sheet (defined below) -> use opening row
        put(cw, col, C_["intinc"], f"={ASSUMP}!{col}{A['yield']}*{col}__TOPEN__", color=GREEN)  # placeholder, fix later
        put(cw, col, C_["fc"], f"=-'{DEBT}'!{col}{D['tot_fc']}", color=DARKR)
        put(cw, col, C_["nonop"], f"={ASSUMP}!{col}{A['nonopinc']}")
        put(cw, col, C_["pbt"], f"={col}{C_['ebit']}+{col}{C_['intinc']}+{col}{C_['fc']}+{col}{C_['nonop']}", bold=True)
        put(cw, col, C_["tax"], f"=-{col}{C_['pbt']}*{ASSUMP}!{col}{A['tax']}")
        put(cw, col, C_["pat"], f"={col}{C_['pbt']}+{col}{C_['tax']}", bold=True)

r += 1
section(cw, r, "B. FREE CASH FLOW"); r += 1
crow("cf_ebitda", "EBITDA")
crow("cf_tax", "(-) Cash tax")
crow("cf_wc", "(-) Increase in working capital")
crow("cfo", "Operating cash flow", bold=True)
crow("cf_capex", "(-) Capex")
crow("cf_leaseprin", "(-) Lease principal repayment")
crow("cf_leaseint", "(-) Lease interest paid")
crow("cf_othint", "(-) Other interest paid")
crow("cf_borrow", "(+/-) Net borrowings draw/(repay)")
crow("fcfe_predivd", "Cash flow before dividends & interest income", bold=True)
crow("cf_intinc", "(+) Interest income received", color=GREEN)
crow("cf_divd", "(-) Dividends paid")
crow("netchg", "Net change in treasury", bold=True)

for i, col in enumerate(COLS):
    put(cw, col, C_["cf_ebitda"], f"={col}{C_['ebitda']}")
    put(cw, col, C_["cf_tax"], f"={col}{C_['tax']}")  # tax already negative
    # working capital change (forecast only)
    if col in HCOLS or i == 0:
        put(cw, col, C_["cf_wc"], None)
    else:
        put(cw, col, C_["cf_wc"], f"=-({col}{C_['ns']}-{COLS[i-1]}{C_['ns']})*{ASSUMP}!{col}{A['wcpct']}")
    put(cw, col, C_["cfo"], f"={col}{C_['cf_ebitda']}+{col}{C_['cf_tax']}+IF({col}{C_['cf_wc']}=\"\",0,{col}{C_['cf_wc']})", bold=True)
    put(cw, col, C_["cf_capex"], f"=-{ASSUMP}!{col}{A['capex']}")
    put(cw, col, C_["cf_leaseprin"], f"=-{ASSUMP}!{col}{A['leaserepay']}")
    put(cw, col, C_["cf_leaseint"], f"=-'{DEBT}'!{col}{D['tot_lease_int']}")
    put(cw, col, C_["cf_othint"], f"=-'{DEBT}'!{col}{D['tot_oth_int']}")
    put(cw, col, C_["cf_borrow"], f"={ASSUMP}!{col}{A['borrowdraw']}-{ASSUMP}!{col}{A['borrowrepay']}")
    put(cw, col, C_["fcfe_predivd"], f"={col}{C_['cfo']}+{col}{C_['cf_capex']}+{col}{C_['cf_leaseprin']}+{col}{C_['cf_leaseint']}+{col}{C_['cf_othint']}+{col}{C_['cf_borrow']}", bold=True)
    # interest income received
    if col in HCOLS:
        put(cw, col, C_["cf_intinc"], None, color=GREEN)
        put(cw, col, C_["cf_divd"], None)
    else:
        put(cw, col, C_["cf_intinc"], f"={col}{C_['intinc']}", color=GREEN)
        put(cw, col, C_["cf_divd"], f"=-{col}{C_['pat']}*{ASSUMP}!{col}{A['payout']}")
    if col in HCOLS:
        put(cw, col, C_["netchg"], None, bold=True)
    else:
        put(cw, col, C_["netchg"], f"={col}{C_['fcfe_predivd']}+{col}{C_['cf_intinc']}+{col}{C_['cf_divd']}", bold=True)

r += 1
section(cw, r, "C. TREASURY ROLL-FORWARD (cash + deposits + MF)"); r += 1
crow("t_open", "Opening treasury")
crow("t_chg", "(+) Net change in treasury")
crow("t_close", "Closing treasury", bold=True, color=GREEN)
crow("t_intinc_chk", "Memo: interest income (yield x opening)", color=GREEN, indent=2)
for i, col in enumerate(COLS):
    if col in HCOLS:
        put(cw, col, C_["t_open"], (f"='{HIST}'!{COLS[i-1]}{H['treas']}" if i>0 else None))
        put(cw, col, C_["t_chg"], None)
        put(cw, col, C_["t_close"], f"='{HIST}'!{col}{H['treas']}", bold=True, color=GREEN)
    else:
        put(cw, col, C_["t_open"], f"={COLS[i-1]}{C_['t_close']}")
        put(cw, col, C_["t_chg"], f"={col}{C_['netchg']}")
        put(cw, col, C_["t_close"], f"={col}{C_['t_open']}+{col}{C_['t_chg']}", bold=True, color=GREEN)
    put(cw, col, C_["t_intinc_chk"], f"={ASSUMP}!{col}{A['yield']}*{col}{C_['t_open']}", color=GREEN, indent=2)

# Now fix the interest income placeholder __TOPEN__ -> opening treasury row
for col in FCOLS:
    cell = cw[f"{col}{C_['intinc']}"]
    cell.value = f"={ASSUMP}!{col}{A['yield']}*{col}{C_['t_open']}"

CASH = "Cash Schedule"

# =====================================================================
# SHEET 6: OUTPUT SUMMARY
# =====================================================================
ow = wb.create_sheet("Output Summary")
ow.sheet_view.showGridLines = False
set_widths(ow, label_w=46, data_w=11)
ow.merge_cells("A1:K1")
style_title(ow, "A1", "  OUTPUT SUMMARY  -  Debt, Net Debt, Interest & Ratios   (Rs crore)")
HR = 3
write_year_header(ow, HR)
O = {}
r = HR + 1
def orow(key, text, fmt=NUMFMT, bold=False, indent=1, color=None):
    global r
    label(ow, r, text, bold=bold, indent=indent, color=(color or "000000"))
    O[key] = r
    r += 1
    return O[key]

section(ow, r, "DEBT & LIQUIDITY"); r += 1
orow("leases", "Total lease liabilities", bold=True)
orow("borrow", "Total borrowings")
orow("debt", "Total debt", bold=True, color=DARKR)
orow("treas", "Treasury (cash + deposits + MF)", bold=True, color=GREEN)
orow("netdebt", "Net debt / (net cash)", bold=True)
for col in COLS:
    put(ow, col, O["leases"], f"='{DEBT}'!{col}{D['l_close']}", bold=True)
    put(ow, col, O["borrow"], f"='{DEBT}'!{col}{D['b_close']}")
    put(ow, col, O["debt"], f"='{DEBT}'!{col}{D['tot_debt']}", bold=True, color=DARKR)
    put(ow, col, O["treas"], f"='{CASH}'!{col}{C_['t_close']}", bold=True, color=GREEN)
    put(ow, col, O["netdebt"], f"={col}{O['debt']}-{col}{O['treas']}", bold=True)

r += 1
section(ow, r, "INTEREST"); r += 1
orow("intexp", "Interest expense (finance cost)", color=DARKR)
orow("intinc", "Interest income", color=GREEN)
orow("netint", "Net interest expense / (income)", bold=True)
for i, col in enumerate(COLS):
    put(ow, col, O["intexp"], f"='{DEBT}'!{col}{D['tot_fc']}", color=DARKR)
    if col in HCOLS:
        put(ow, col, O["intinc"], None, color=GREEN)
        put(ow, col, O["netint"], None, bold=True)
    else:
        put(ow, col, O["intinc"], f"='{CASH}'!{col}{C_['intinc']}", color=GREEN)
        put(ow, col, O["netint"], f"={col}{O['intexp']}-{col}{O['intinc']}", bold=True)

r += 1
section(ow, r, "KEY RATIOS"); r += 1
orow("ebitda", "EBITDA")
orow("nd_ebitda", "Net debt / EBITDA (x)", fmt='0.00"x"')
orow("icr", "EBITDA / interest expense (x)", fmt='0.0"x"')
orow("gd_eq", "Gross debt / equity (x)", fmt='0.00"x"')
for col in COLS:
    put(ow, col, O["ebitda"], f"='{CASH}'!{col}{C_['ebitda']}")
    put(ow, col, O["nd_ebitda"], f"={col}{O['netdebt']}/{col}{O['ebitda']}", fmt='0.00"x"')
    put(ow, col, O["icr"], f"=IF({col}{O['intexp']}=0,0,{col}{O['ebitda']}/{col}{O['intexp']})", fmt='0.0"x"')
# equity for gd/eq (hist + forecast retained earnings build)
# build equity forecast: equity_t = equity_{t-1} + PAT - dividends
eq_row = O["gd_eq"]
# need an equity line; add memo
r += 1
section(ow, r, "MEMO: EQUITY ROLL"); r += 1
orow("eq_open", "Opening equity")
orow("eq_pat", "(+) PAT")
orow("eq_div", "(-) Dividends")
orow("eq_close", "Closing equity", bold=True)
for i, col in enumerate(COLS):
    if col in HCOLS:
        put(ow, col, O["eq_open"], (f"='{HIST}'!{COLS[i-1]}{H['equity']}" if i>0 else None))
        put(ow, col, O["eq_pat"], f"='{CASH}'!{col}{C_['pat']}")
        put(ow, col, O["eq_div"], None)
        put(ow, col, O["eq_close"], f"='{HIST}'!{col}{H['equity']}", bold=True)
    else:
        put(ow, col, O["eq_open"], f"={COLS[i-1]}{O['eq_close']}")
        put(ow, col, O["eq_pat"], f"='{CASH}'!{col}{C_['pat']}")
        put(ow, col, O["eq_div"], f"=-'{CASH}'!{col}{C_['cf_divd']}*-1")  # dividends paid (positive shown negative)
        put(ow, col, O["eq_div"], f"='{CASH}'!{col}{C_['cf_divd']}")
        put(ow, col, O["eq_close"], f"={col}{O['eq_open']}+{col}{O['eq_pat']}+{col}{O['eq_div']}", bold=True)
# now gd/eq
for col in COLS:
    put(ow, col, O["gd_eq"], f"={col}{O['debt']}/{col}{O['eq_close']}", fmt='0.00"x"')

# Freeze panes & tab colors
for ws in [aw, hw, dw, cw, ow]:
    ws.freeze_panes = "C4"
cover.sheet_properties.tabColor = NAVY
aw.sheet_properties.tabColor = GOLD
hw.sheet_properties.tabColor = BLUE
dw.sheet_properties.tabColor = DARKR
cw.sheet_properties.tabColor = GREEN
ow.sheet_properties.tabColor = NAVY

# Enable iterative calc just in case (safe)
try:
    wb.calculation.iterate = True
    wb.calculation.iterateCount = 100
    wb.calculation.iterateDelta = 0.001
    wb.calculation.fullCalcOnLoad = True
except Exception as e:
    print("calc props:", e)

wb.save("USL_Debt_Cash_Schedule.xlsx")
print("Saved USL_Debt_Cash_Schedule.xlsx")
