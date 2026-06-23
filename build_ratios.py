import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "/projects/sandbox/USLL/United Spirits Excell - 8 Year Forecast (Dep Corrected).xlsx"
OUT = "/projects/sandbox/USLL/United Spirits Excell - 8 Year + Ratios.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=False)
if "Ratios" in wb.sheetnames:
    del wb["Ratios"]
ws = wb.create_sheet("Ratios")

COLS = [get_column_letter(i) for i in range(3, 21)]   # C..T = 2017..2034

# ---- styles ----
hdr   = Font(bold=True, size=12)
sect  = Font(bold=True, color="1F4E78")
yrfnt = Font(bold=True)
fill  = PatternFill("solid", fgColor="DDEBF7")
ctr   = Alignment(horizontal="center")
bottom= Border(bottom=Side(style="thin"))

ws["B2"] = "Ratio Analysis  (FY2017 - FY2034)"; ws["B2"].font = hdr
ws["B4"] = "Particulars"; ws["B4"].font = yrfnt
ws["C4"] = 2017
for k in range(1, len(COLS)):
    ws[f"{COLS[k]}4"] = f"={COLS[k-1]}4+1"
for c in COLS:
    ws[f"{c}4"].font = yrfnt; ws[f"{c}4"].alignment = ctr; ws[f"{c}4"].border = bottom

PCT = "0.0%"; R2 = "0.00"; RX = '0.00"x"'; CX = '0.0"x"'; D0 = "0"; NUM = "#,##0"
rows = [
    (6,  "Profitability Ratios", None, None),
    (7,  "Gross Margin",                       "=PL!{X}8/PL!{X}6",                          PCT),
    (8,  "EBITDA Margin",                      "=PL!{X}13/PL!{X}6",                         PCT),
    (9,  "EBIT Margin",                        "=PL!{X}15/PL!{X}6",                         PCT),
    (10, "Net Profit Margin",                  "=PL!{X}23/PL!{X}6",                         PCT),
    (11, "Return on Equity (ROE)",             "=PL!{X}23/BS!{X}12",                        PCT),
    (12, "Return on Capital Employed (ROCE)",  "=PL!{X}15/(BS!{X}48-BS!{X}25)",             PCT),
    (14, "Liquidity Ratios", None, None),
    (15, "Current Ratio",                      "=BS!{X}46/BS!{X}25",                        R2),
    (16, "Quick Ratio",                        "=(BS!{X}46-BS!{X}43)/BS!{X}25",             R2),
    (18, "Leverage Ratios", None, None),
    (19, "Debt Equity Ratio",                  "=(BS!{X}15+BS!{X}17+BS!{X}22)/BS!{X}12",    R2),
    (20, "Net Debt to EBITDA",                 "=((BS!{X}15+BS!{X}17+BS!{X}22)-BS!{X}45)/PL!{X}13", R2),
    (21, "Interest Coverage Ratio",            "=PL!{X}15/-PL!{X}18",                       CX),
    (23, "Efficiency Ratios", None, None),
    (24, "Asset Turnover Ratio",               "=PL!{X}6/BS!{X}48",                         RX),
    (25, "Fixed Asset Turnover Ratio",         "=PL!{X}6/BS!{X}37",                         RX),
    (27, "Working Capital Ratios", None, None),
    (28, "Receivables days",                   "='Working Capital Schedule'!{X}10",         D0),
    (29, "Payables days",                      "='Working Capital Schedule'!{X}16",         D0),
    (30, "Inventory days",                     "='Working Capital Schedule'!{X}13",         D0),
    (32, "Valuation Ratios", None, None),
    (33, "EV/EBITDA",                          "={X}46/PL!{X}13",                           CX),
    (34, "EV/Sales",                           "={X}46/PL!{X}6",                            CX),
    (35, "EV/EBIT",                            "={X}46/PL!{X}15",                           CX),
    (36, "PE Ratio",                           "={X}42/PL!{X}23",                           CX),
    (39, "Memo - valuation basis  (actual year-end market price; FY27-34 held at current price Rs1,246.7)", None, None),
    (42, "Market Capitalisation (Rs cr)",      "={X}40*{X}41",                              NUM),
    (43, "Total Debt (borrowings + leases)",   "=BS!{X}15+BS!{X}17+BS!{X}22",               NUM),
    (44, "Cash & Investments",                 "=BS!{X}45",                                 NUM),
    (45, "Net Debt",                           "={X}43-{X}44",                              NUM),
    (46, "Enterprise Value (Mkt Cap + Net Debt)","={X}42+{X}45",                            NUM),
]

# --- actual year-end share prices (Rs): FY17-19 implied from annual-report market cap;
#     FY20 = COVID-trough estimate; FY21-26 = model Beta sheet; FY27-34 = current price held flat ---
PRICE = {"C":434.3, "D":624.5, "E":539.5, "F":490.0, "G":519.5, "H":862.9, "I":777.1,
         "J":1128.3, "K":1427.3, "L":1325.6}
for fc in ["M","N","O","P","Q","R","S","T"]:
    PRICE[fc] = 1246.7

for row, label, tmpl, fmt in rows:
    ws[f"B{row}"] = label
    if tmpl is None:                       # section header
        ws[f"B{row}"].font = sect
        for c in COLS + ["B"]:
            ws[f"{c}{row}"].fill = fill
        continue
    for c in COLS:
        cell = ws[f"{c}{row}"]
        cell.value = tmpl.format(X=c)
        cell.number_format = fmt

# memo: year-end share price (row 40, per-year actuals) and shares (row 41)
ws["B40"] = "Year-end Share Price (Rs)"
ws["B41"] = "Shares Outstanding (cr)"
for c in COLS:
    pc = ws[f"{c}40"]; pc.value = PRICE[c]; pc.number_format = "#,##0.0"
    sc = ws[f"{c}41"]; sc.value = 72.74;    sc.number_format = "0.00"

# column widths
ws.column_dimensions["B"].width = 34
for c in COLS:
    ws.column_dimensions[c].width = 8.5
ws.freeze_panes = "C5"
wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("SAVED", OUT, "| sheets:", len(wb.sheetnames))
