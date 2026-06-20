"""Correct the depreciation schedule in the 8-year model:
 (1) PPE  : depreciate the EXISTING block on GROSS cost / useful life (was: net book value
            over a fresh full life, which understated the charge). Capex vintage layers unchanged.
 (2) ROU  : replace the 7-year vintage waterfall (which over-runs a sub-horizon life) with the
            standard % method already defined in Assumptions row 16 (30% of opening + 1/2 additions),
            which matches the annual-report effective amortisation rate (~30-35%, ~3-yr life).
Everything stays integrated; the workbook continues to balance."""
import openpyxl
from openpyxl.utils import column_index_from_string as ci
from copy import copy

SRC = "/projects/sandbox/USLL/United Spirits Excell - 8 Year Forecast.xlsx"
OUT = "/projects/sandbox/USLL/United Spirits Excell - 8 Year Forecast (Dep Corrected).xlsx"
wb = openpyxl.load_workbook(SRC, data_only=False)
ws = wb["Working Schedules"]
COLS = ["H","I","J","K","L","M","N","O"]            # forecast 2027..2034
def asmcol(c):                                       # WS col -> Assumptions col (offset 3)
    return chr(ord(c) - 3)
def setc(col, row, val, style_from=None):
    cell = ws.cell(row=row, column=ci(col))
    cell.value = val
    if style_from is not None and style_from.has_style:
        cell._style = copy(style_from._style)
    return cell

# ---------- (1) PPE: existing block on GROSS cost / useful life ----------
# G65 = 2026 closing GROSS PPE (last actual); R86 = useful life (16). Constant charge.
setc("H", 90, "=-$G$65/$R$86", ws["H90"])
for k in range(1, 8):                                # keep copies constant across years
    setc(COLS[k], 90, f"={COLS[k-1]}90", ws[f"{COLS[k]}90"])

# ---------- (2) ROU: switch waterfall -> % method (Assumptions row 16) ----------
sty_lab  = ws["B105"]; sty_num = ws["H107"]; sty_yhdr = ws["H103"]
# clear the old vintage waterfall body rows 105..117 (cols B and H..O)
for r in range(104, 118):
    for col in ["B"] + COLS:
        ws.cell(row=r, column=ci(col)).value = None
# repurpose the useful-life helper to show the amortisation rate used
setc("Q", 103, "Amortisation %", ws["Q103"]); setc("R", 103, "=Assumptions!E16", ws["R103"])
# readable mini-schedule: Opening / Additions / Depreciation / Closing
setc("B", 105, "Opening", sty_lab)
setc("B", 106, "Additions", sty_lab)
setc("B", 107, "Depreciation (30% x [Opening + 1/2 Additions])", sty_lab)
setc("B", 108, "Closing", sty_lab)
for k, c in enumerate(COLS):
    prev = COLS[k-1] if k > 0 else "G"
    a = asmcol(c)
    setc(c, 105, ("=G55" if k == 0 else f"={prev}108"), sty_num)      # opening = prior closing
    setc(c, 106, f"={c}52", sty_num)                                  # additions (ROU asset sched)
    setc(c, 107, f"=-Assumptions!{a}16*({c}105+0.5*{c}106)", sty_num) # % depreciation
    setc(c, 108, f"={c}105+{c}106+{c}107", sty_num)                   # closing

# ---------- redirect downstream references from old totals to new ROU dep row 107 ----------
for k, c in enumerate(COLS):
    prev = COLS[k-1] if k > 0 else "G"
    setc(c, 53, f"={c}107", ws[f"{c}53"])                 # ROU asset-schedule depreciation
    setc(c, 70, f"=-{c}82+{c}99+{c}107", ws[f"{c}70"])    # Total D&A (PPE row99 + ROU row107 + amort)

wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("SAVED", OUT)
