"""
Build a LIVE, FORMULA-DRIVEN, fully-linked Excel workbook: USL_Financial_Model.xlsx
Colour convention (standard financial-modelling):
  BLUE  (#0000FF) = hard-coded inputs (historical actuals + forecast assumptions)
  BLACK (#000000) = calculations (formulas, incl. those combining other-sheet cells)
  GREEN (#008000) = direct links / references to another worksheet (= 'Sheet'!Cell)
Non-circular by design (treasury income uses opening balance); the 3 statements are
linked and the balance sheet balances every year; the cash flow ties to the change in
treasury. Reads engine values for hard-coded cells; everything else is an Excel formula.
"""
import re, pickle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

st = pickle.load(open("build_state.pkl","rb"))
M=st["M"]; A=st["A"]
HIST=["FY22","FY23","FY24","FY25","FY26"]; FCST=["FY27E","FY28E","FY29E","FY30E"]; YEARS=HIST+FCST
COLS=["B","C","D","E","F","G","H","I","J"]; YC=dict(zip(YEARS,COLS)); HC=COLS[:5]; FCC=COLS[5:]
SHARES=72.78

# ---------- styles ----------
BLUE="0000FF"; BLACK="000000"; GREEN="008000"; NAVY="1F3864"; HDRB="2E5496"; LBLUE="D9E1F2"; GREY="F2F2F2"; GOLD="FFF2CC"; GREENF="E2EFDA"
fillN=PatternFill("solid",fgColor=NAVY); fillH=PatternFill("solid",fgColor=HDRB)
fillL=PatternFill("solid",fgColor=LBLUE); fillG=PatternFill("solid",fgColor=GREY); fillGo=PatternFill("solid",fgColor=GOLD)
thin=Side(style="thin",color="D0D0D0"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
RGT=Alignment(horizontal="right"); LFT=Alignment(horizontal="left",wrap_text=True); CTR=Alignment(horizontal="center",wrap_text=True)
NUM="#,##0;(#,##0)"; NUM1="#,##0.0;(#,##0.0)"; PC="0.0%"; XF='0.0"x"'; RUP='"\u20b9"#,##0'
LINK=re.compile(r"^=-?'[^']+'!\$?[A-Z]{1,3}\$?\d+$")  # pure cross-sheet link -> green

wb=Workbook(); RM={}

def fcolor(v):
    if isinstance(v,str) and v.startswith("="):
        return GREEN if LINK.match(v) else BLACK
    return BLUE

def cset(ws,row,colletter,val,fmt=None,bold=False,fill=None,color=None,align=RGT):
    c=ws.cell(row=row,column=column_index_from_string(colletter),value=val)
    c.font=Font(color=(color or fcolor(val)),bold=bold,size=10)
    if fmt:c.number_format=fmt
    if fill:c.fill=fill
    c.alignment=align; c.border=border
    return c

def label(ws,row,text,bold=False,fill=None,indent=0):
    c=ws.cell(row=row,column=1,value=("   "*indent)+text)
    c.font=Font(bold=bold,size=10,color=BLACK)
    if fill:c.fill=fill
    c.alignment=LFT; c.border=border
    return c

def header(ws,title,years=YEARS,ncol=10):
    ws.sheet_view.showGridLines=False
    c=ws.cell(row=1,column=1,value=title); c.font=Font(bold=True,color="FFFFFF",size=12); c.fill=fillH; c.alignment=LFT
    for j in range(2,ncol+2): ws.cell(row=1,column=j).fill=fillH
    cc=ws.cell(row=2,column=1,value="\u20b9 crore (unless stated)"); cc.font=Font(bold=True,color="FFFFFF"); cc.fill=fillN; cc.alignment=LFT; cc.border=border
    for j,y in enumerate(years):
        h=ws.cell(row=2,column=2+j,value=y); h.font=Font(bold=True,color="FFFFFF"); h.fill=fillN; h.alignment=CTR; h.border=border
    ws.column_dimensions["A"].width=42
    for j in range(len(years)): ws.column_dimensions[get_column_letter(2+j)].width=12

# ---------- PASS 1: allocate rows for every keyed line on every sheet ----------
LAYOUT={
 "Assumptions":[("__t__","FORECAST DRIVERS (BASE CASE) — blue = analyst inputs"),
   ("vol_growth","Volume growth %"),("real_growth","Realisation (NSV/case) growth %"),
   ("bev_ebitda_margin","Core EBITDA margin (% NSV)"),("gross_margin_nsv","Gross margin (% NSV)"),
   ("ad_promo_pct_nsv","A&P (% NSV)"),("employee_pct_nsv","Employee (% NSV)"),
   ("excise_pct_gross","Excise (% gross revenue)"),("tax_rate","Effective tax rate %"),
   ("payout","Dividend payout %"),("treasury_yield","Treasury yield %"),
   ("capex_pct_nsv","Capex (% NSV)"),("ppe_dep_rate","PPE depreciation rate %"),
   ("lease_add_pct_nsv","New leases (% NSV)"),("rou_dep_rate","ROU depreciation rate %"),
   ("lease_int_rate","Lease interest rate %"),("lease_pay_pct","Lease payment (% opening) %"),
   ("recv_days_gross","Receivable days (gross rev)"),("inv_days_cogs","Inventory days (COGS)"),
   ("pay_days_cogs","Payable days (COGS)"),("recv_nc_pct_nsv","Non-cur receivables (% NSV)"),
   ("other_cl_pct_nsv","Other current liab (% NSV)"),("provisions_pct_nsv","Provisions (% NSV)"),
   ("finance_cost_other","Other finance cost (\u20b9 cr)"),
   ("__t2__","MODEL ANCHORS & CONSTANTS  (value in column B)"),
   ("shares","Shares outstanding (cr)"),("ppe_owned_base","FY26 owned PPE (\u20b9 cr)"),
   ("rou_base","FY26 ROU asset (\u20b9 cr)"),("intangible_amort","Intangible amortisation /yr"),
   ("treasury_const","Treasury other income constant"),("cash_split","Cash split of treasury"),
   ("dep_split","Deposit split of treasury"),("inv_split","Investment split of treasury"),
   ("tax_protest","Tax paid under protest (held flat)"),("other_assets_base","Other op. assets base (FY26)"),
   ("curtax_growth","Current tax liab growth"),("rcb_value","RCB optionality value (\u20b9 cr)"),
   ("wacc","WACC"),("term_g","Terminal growth g")],
 "Revenue Buildup":[("volume","Volume (mn cases)"),("real","NSV per case (\u20b9) - realisation"),
   ("nsv","Net Sales Value (NSV) = Vol x Realisation"),("excise","Excise duty"),
   ("gross","Gross revenue (incl. excise)"),("pa_nsv","P&A salience (% of NSV)"),("pa_vol","P&A salience (% of volume)")],
 "Income Statement":[("nsv","Net Sales Value (NSV)"),("cogs","Cost of goods sold"),("gp","Gross Profit"),
   ("employee","Employee cost"),("ad_promo","Advertising & promotion"),("opex","Other operating expense"),
   ("ebitda","EBITDA (core)"),("dep","Depreciation & amortisation"),("ebit","EBIT"),
   ("oi","Other income (treasury)"),("fin","Finance costs"),("exc","Exceptional / JV"),
   ("pbt","Profit before tax"),("tax","Tax"),("patc","PAT (continuing)"),("disc","Discontinued ops (RCB)"),
   ("pat","PAT (reported)"),("eps","EPS (\u20b9)")],
 "Working Capital":[("recv","Trade receivables (current)"),("inv","Inventories"),("pay","Trade payables"),
   ("recvnc","Trade receivables (non-current)"),("nwc","Net trade working capital"),
   ("recvd","Receivable days (gross rev)"),("invd","Inventory days (COGS)"),("payd","Payable days (COGS)"),
   ("ccc","Cash conversion cycle (days)")],
 "PPE Schedule":[("__s1__","Owned PPE"),("ppe_open","  Opening (net)"),("capex","  + Capex"),
   ("ppe_dep","  - Depreciation"),("ppe_close","  Closing owned PPE (net)"),
   ("__s2__","Right-of-use (ROU) asset"),("rou_open","  Opening"),("rou_add","  + New leases"),
   ("rou_dep","  - ROU depreciation"),("rou_close","  Closing ROU"),
   ("ppe_rep","Reported Net PPE (owned + ROU)"),("cwip","Capital work-in-progress"),
   ("capex_s","Capex / NSV %"),("turn","Asset turnover (NSV/Net PPE)")],
 "Lease Schedule":[("open","Lease liability - opening"),("new","  + New leases"),("int","  + Interest accretion"),
   ("pay","  - Lease payments"),("close","Lease liability - closing"),("rou","ROU asset (memo)"),
   ("intpl","Lease interest in P&L")],
 "Debt & Cash":[("borrow","Borrowings (gross debt)"),("lease","Lease liabilities"),("totdebt","Total debt (incl leases)"),
   ("cash","Cash & equivalents"),("dep","Bank deposits"),("inv","Current investments"),
   ("treas","Total cash & treasury"),("netdebt","Net debt / (net cash)")],
 "Balance Sheet":[("__a__","ASSETS"),("ppe","Net PPE (incl ROU)"),("cwip","Capital work-in-progress"),
   ("gw","Goodwill"),("intang","Other intangible assets"),("invprop","Investment property"),
   ("jv","Investment in JVs"),("inv","Inventories"),("recv","Trade receivables (current)"),
   ("recvnc","Trade receivables (non-current)"),("ofa","Other financial assets"),("dta","Deferred tax assets"),
   ("cta","Current tax assets"),("oa","Other assets (incl tax under protest)"),("hfsa","Assets held for sale (RCB)"),
   ("treas","Cash + deposits + investments (treasury, PLUG)"),("nca","  Non-cash assets subtotal"),
   ("ta","TOTAL ASSETS"),
   ("__e__","EQUITY & LIABILITIES"),("eq","Total equity"),("div","  Memo: dividend paid"),
   ("borrow","Borrowings"),("lease","Lease liabilities"),("pay","Trade payables"),("dtl","Deferred tax liabilities"),
   ("prov","Provisions"),("ctl","Current tax liabilities"),("ofl","Other financial liabilities"),
   ("ocl","Other current liabilities"),("hfsl","Liabilities of disposal group (RCB)"),
   ("tl","  Total liabilities (excl equity)"),("tle","TOTAL EQUITY & LIABILITIES"),
   ("chk","Balance check (Assets - E&L)")],
 "Cash Flow":[("pat","PAT (reported)"),("da","+ Depreciation & amortisation"),("lint","+ Lease interest (add back)"),
   ("dwc","+/- Change in working capital"),("cfo","Operating cash flow (CFO)"),
   ("capex","- Capex"),("rcb","+ RCB / held-for-sale disposal"),("cfi","Investing cash flow (CFI)"),
   ("divp","- Dividends paid"),("leasep","- Lease payments"),("dborrow","+/- Change in borrowings"),
   ("cff","Financing cash flow (CFF)"),("net","Net change in cash & treasury"),
   ("topen","Opening treasury"),("tclose","Closing treasury (= opening + net)"),("tie","Tie check vs Balance Sheet")],
}

dcf_block=None  # DCF/Sens/Comps/HistVal/Scenarios laid out manually later

def alloc(name, lines, years=YEARS, ncol=10, sources=False):
    ws=wb.create_sheet(name); RM[name]={}
    header(ws,name if name!="Assumptions" else "ASSUMPTIONS — drivers, anchors & constants",years,ncol)
    r=3
    for key,lab in lines:
        if key.startswith("__"):
            label(ws,r,lab,bold=True,fill=fillL)
            for j in range(len(years)): ws.cell(row=r,column=2+j).fill=fillL
            r+=1; continue
        bold = lab.isupper() or key in ("nsv","gp","ebitda","ebit","pbt","pat","ppe_close","rou_close","ppe_rep",
                                        "close","totdebt","treas","netdebt","ta","tle","eq","cfo","cfi","cff","net","ccc","nwc")
        label(ws,r,lab,bold=bold)
        RM[name][key]=r; r+=1
    return ws

WS={}
for nm,lines in LAYOUT.items():
    WS[nm]=alloc(nm,lines)

# ---------- helpers for references ----------
def RS(sheet,key,year): return f"'{sheet}'!{YC[year]}{RM[sheet][key]}"          # cross-sheet (green if alone)
def RSa(sheet,key):     return f"'{sheet}'!$B${RM[sheet][key]}"                  # anchor (col B abs)
def SAME(sheet,key,col):return f"{col}{RM[sheet][key]}"                          # same-sheet (black)
def prevcol(col): return COLS[COLS.index(col)-1]

# ===================================================================================
# PASS 2 — FILL CELLS
# ===================================================================================
# ---- Assumptions ----
ws=WS["Assumptions"]
drv=["vol_growth","real_growth","bev_ebitda_margin","gross_margin_nsv","ad_promo_pct_nsv","employee_pct_nsv",
     "excise_pct_gross","tax_rate","payout","treasury_yield","capex_pct_nsv","ppe_dep_rate","lease_add_pct_nsv",
     "rou_dep_rate","lease_int_rate","lease_pay_pct","recv_days_gross","inv_days_cogs","pay_days_cogs",
     "recv_nc_pct_nsv","other_cl_pct_nsv","provisions_pct_nsv","finance_cost_other"]
daycnt={"recv_days_gross","inv_days_cogs","pay_days_cogs","finance_cost_other"}
for k in drv:
    rr=RM["Assumptions"][k]
    for y in FCST:
        fmt = NUM1 if k in daycnt else PC
        cset(ws,rr,YC[y],round(A[k][y],4),fmt)
anchors={"shares":SHARES,"ppe_owned_base":900,"rou_base":315,"intangible_amort":15,"treasury_const":60,
 "cash_split":0.27,"dep_split":0.36,"inv_split":0.37,"tax_protest":1191,"other_assets_base":667,
 "curtax_growth":1.05,"rcb_value":2500,"wacc":0.1025,"term_g":0.0625}
for k,v in anchors.items():
    fmt = PC if k in ("wacc","term_g") else (NUM1 if k in("shares","curtax_growth") else NUM)
    cset(ws,RM["Assumptions"][k],"B",v,fmt)
cset(ws,2,"A","\u20b9 crore / % (blue = hard-coded analyst input)",color="FFFFFF",align=LFT,fill=fillN,bold=True)

# ---- Revenue Buildup ----
S="Revenue Buildup"; ws=WS[S]
for y in HIST:
    cset(ws,RM[S]["volume"],YC[y],M["volume_mn_cases"][y],NUM1)
    cset(ws,RM[S]["real"],YC[y],round(M["nsv_per_case"][y]),NUM)
    cset(ws,RM[S]["nsv"],YC[y],round(M["net_revenue"][y],1),NUM,bold=True)
    cset(ws,RM[S]["excise"],YC[y],round(M["excise_duty"][y]),NUM)
    cset(ws,RM[S]["gross"],YC[y],round(M["gross_revenue"][y]),NUM)
    cset(ws,RM[S]["pa_nsv"],YC[y],M["pa_nsv_salience"][y],NUM)
    cset(ws,RM[S]["pa_vol"],YC[y],M["pa_vol_salience"][y],NUM)
for y in FCST:
    c=YC[y]; pc=prevcol(c)
    cset(ws,RM[S]["volume"],c,f"={SAME(S,'volume',pc)}*(1+{RS('Assumptions','vol_growth',y)})",NUM1)
    cset(ws,RM[S]["real"],c,f"={SAME(S,'real',pc)}*(1+{RS('Assumptions','real_growth',y)})",NUM)
    cset(ws,RM[S]["nsv"],c,f"={SAME(S,'volume',c)}*1000000*{SAME(S,'real',c)}/10000000",NUM,bold=True)
    cset(ws,RM[S]["excise"],c,f"={SAME(S,'nsv',c)}/(1-{RS('Assumptions','excise_pct_gross',y)})*{RS('Assumptions','excise_pct_gross',y)}",NUM)
    cset(ws,RM[S]["gross"],c,f"={SAME(S,'nsv',c)}+{SAME(S,'excise',c)}",NUM)
    cset(ws,RM[S]["pa_nsv"],c,f"=MIN(95,{SAME(S,'pa_nsv',pc)}+0.8)",NUM)
    cset(ws,RM[S]["pa_vol"],c,f"=MIN(92,{SAME(S,'pa_vol',pc)}+1.2)",NUM)

# ---- Income Statement ----
S="Income Statement"; ws=WS[S]
for y in YEARS:  # NSV always linked from Revenue Buildup
    cset(ws,RM[S]["nsv"],YC[y],f"={RS('Revenue Buildup','nsv',y)}",NUM,bold=True)
for y in HIST:
    for k,src in [("cogs","cogs"),("gp","bev_gp"),("employee","employee"),("ad_promo","ad_promo"),
                  ("opex","other_expense"),("ebitda","bev_ebitda"),("dep","dep_amort"),("oi","other_income"),
                  ("fin","finance_cost"),("patc","pat_continuing"),("disc","disc_ops_pat"),
                  ("pat","pat_reported"),("tax","tax")]:
        cset(ws,RM[S][k],YC[y],round(M[src][y],1),NUM,bold=(k in("gp","ebitda","pat")))
    cset(ws,RM[S]["exc"],YC[y],round(M["exceptional"][y]+M["jv_share"][y],1),NUM)
    cset(ws,RM[S]["pbt"],YC[y],round(M["pbt"][y],1),NUM,bold=True)
    cset(ws,RM[S]["eps"],YC[y],round(M["eps_reported"][y],2),NUM1)
for y in YEARS:  # EBIT computed everywhere (black)
    c=YC[y]; cset(ws,RM[S]["ebit"],c,f"={SAME(S,'ebitda',c)}-{SAME(S,'dep',c)}",NUM,bold=True)
for y in FCST:
    c=YC[y]; pc=prevcol(c)
    cset(ws,RM[S]["gp"],c,f"={SAME(S,'nsv',c)}*{RS('Assumptions','gross_margin_nsv',y)}",NUM,bold=True)
    cset(ws,RM[S]["cogs"],c,f"={SAME(S,'nsv',c)}-{SAME(S,'gp',c)}",NUM)
    cset(ws,RM[S]["employee"],c,f"={SAME(S,'nsv',c)}*{RS('Assumptions','employee_pct_nsv',y)}",NUM)
    cset(ws,RM[S]["ad_promo"],c,f"={SAME(S,'nsv',c)}*{RS('Assumptions','ad_promo_pct_nsv',y)}",NUM)
    cset(ws,RM[S]["ebitda"],c,f"={SAME(S,'nsv',c)}*{RS('Assumptions','bev_ebitda_margin',y)}",NUM,bold=True)
    cset(ws,RM[S]["opex"],c,f"={SAME(S,'gp',c)}-{SAME(S,'employee',c)}-{SAME(S,'ad_promo',c)}-{SAME(S,'ebitda',c)}",NUM)
    cset(ws,RM[S]["dep"],c,f"={RS('PPE Schedule','ppe_dep',y)}+{RS('PPE Schedule','rou_dep',y)}+{RSa('Assumptions','intangible_amort')}",NUM)
    cset(ws,RM[S]["oi"],c,f"={RS('Balance Sheet','treas',YEARS[YEARS.index(y)-1])}*{RS('Assumptions','treasury_yield',y)}+{RSa('Assumptions','treasury_const')}",NUM)
    cset(ws,RM[S]["fin"],c,f"={RS('Lease Schedule','int',y)}+{RS('Assumptions','finance_cost_other',y)}",NUM)
    cset(ws,RM[S]["exc"],c,0,NUM)
    cset(ws,RM[S]["pbt"],c,f"={SAME(S,'ebit',c)}+{SAME(S,'oi',c)}-{SAME(S,'fin',c)}+{SAME(S,'exc',c)}",NUM,bold=True)
    cset(ws,RM[S]["tax"],c,f"={SAME(S,'pbt',c)}*{RS('Assumptions','tax_rate',y)}",NUM)
    cset(ws,RM[S]["patc"],c,f"={SAME(S,'pbt',c)}-{SAME(S,'tax',c)}",NUM)
    cset(ws,RM[S]["disc"],c,0,NUM)
    cset(ws,RM[S]["pat"],c,f"={SAME(S,'patc',c)}+{SAME(S,'disc',c)}",NUM,bold=True)
    cset(ws,RM[S]["eps"],c,f"={SAME(S,'pat',c)}/{RSa('Assumptions','shares')}",NUM1)

# ---- Working Capital ----
S="Working Capital"; ws=WS[S]
for y in HIST:
    cset(ws,RM[S]["recv"],YC[y],round(M["receivables_cur"][y],1),NUM)
    cset(ws,RM[S]["inv"],YC[y],round(M["inventory"][y],1),NUM)
    cset(ws,RM[S]["pay"],YC[y],round(M["trade_payables"][y],1),NUM)
    cset(ws,RM[S]["recvnc"],YC[y],round(M["receivables_nc"][y],1),NUM)
for y in FCST:
    c=YC[y]
    cset(ws,RM[S]["recv"],c,f"={RS('Assumptions','recv_days_gross',y)}/365*{RS('Revenue Buildup','gross',y)}",NUM)
    cset(ws,RM[S]["inv"],c,f"={RS('Assumptions','inv_days_cogs',y)}/365*{RS('Income Statement','cogs',y)}",NUM)
    cset(ws,RM[S]["pay"],c,f"={RS('Assumptions','pay_days_cogs',y)}/365*{RS('Income Statement','cogs',y)}",NUM)
    cset(ws,RM[S]["recvnc"],c,f"={RS('Revenue Buildup','nsv',y)}*{RS('Assumptions','recv_nc_pct_nsv',y)}",NUM)
for y in YEARS:
    c=YC[y]
    cset(ws,RM[S]["nwc"],c,f"={SAME(S,'inv',c)}+{SAME(S,'recv',c)}-{SAME(S,'pay',c)}",NUM,bold=True)
    cset(ws,RM[S]["recvd"],c,f"={SAME(S,'recv',c)}/{RS('Revenue Buildup','gross',y)}*365",NUM1)
    cset(ws,RM[S]["invd"],c,f"={SAME(S,'inv',c)}/{RS('Income Statement','cogs',y)}*365",NUM1)
    cset(ws,RM[S]["payd"],c,f"={SAME(S,'pay',c)}/{RS('Income Statement','cogs',y)}*365",NUM1)
    cset(ws,RM[S]["ccc"],c,f"={SAME(S,'invd',c)}+{SAME(S,'recvd',c)}-{SAME(S,'payd',c)}",NUM1,bold=True)

# ---- PPE Schedule ----
S="PPE Schedule"; ws=WS[S]
for y in HIST:
    cset(ws,RM[S]["capex"],YC[y],M["capex"][y],NUM)
    cset(ws,RM[S]["ppe_rep"],YC[y],round(M["ppe"][y],1),NUM,bold=True)
    cset(ws,RM[S]["cwip"],YC[y],round(M["cwip"][y],1),NUM)
# FY26 anchors for owned PPE / ROU closing
cset(ws,RM[S]["ppe_close"],"F",f"={RSa('Assumptions','ppe_owned_base')}",NUM,bold=True)
cset(ws,RM[S]["rou_close"],"F",f"={RSa('Assumptions','rou_base')}",NUM,bold=True)
for y in FCST:
    c=YC[y]; pc=prevcol(c)
    cset(ws,RM[S]["ppe_open"],c,f"={SAME(S,'ppe_close',pc)}",NUM)
    cset(ws,RM[S]["capex"],c,f"={RS('Revenue Buildup','nsv',y)}*{RS('Assumptions','capex_pct_nsv',y)}",NUM)
    cset(ws,RM[S]["ppe_dep"],c,f"=({SAME(S,'ppe_open',c)}+{SAME(S,'capex',c)}*0.5)*{RS('Assumptions','ppe_dep_rate',y)}",NUM)
    cset(ws,RM[S]["ppe_close"],c,f"={SAME(S,'ppe_open',c)}+{SAME(S,'capex',c)}-{SAME(S,'ppe_dep',c)}",NUM,bold=True)
    cset(ws,RM[S]["rou_open"],c,f"={SAME(S,'rou_close',pc)}",NUM)
    cset(ws,RM[S]["rou_add"],c,f"={RS('Revenue Buildup','nsv',y)}*{RS('Assumptions','lease_add_pct_nsv',y)}",NUM)
    cset(ws,RM[S]["rou_dep"],c,f"=({SAME(S,'rou_open',c)}+{SAME(S,'rou_add',c)}*0.5)*{RS('Assumptions','rou_dep_rate',y)}",NUM)
    cset(ws,RM[S]["rou_close"],c,f"={SAME(S,'rou_open',c)}+{SAME(S,'rou_add',c)}-{SAME(S,'rou_dep',c)}",NUM,bold=True)
    cset(ws,RM[S]["cwip"],c,f"={SAME(S,'cwip',pc)}",NUM)
cset(ws,RM[S]["ppe_rep"],"F",f"={SAME(S,'ppe_close','F')}+{SAME(S,'rou_close','F')}",NUM,bold=True)
for y in FCST:
    c=YC[y]; cset(ws,RM[S]["ppe_rep"],c,f"={SAME(S,'ppe_close',c)}+{SAME(S,'rou_close',c)}",NUM,bold=True)
for y in YEARS:
    c=YC[y]
    cset(ws,RM[S]["capex_s"],c,f"={SAME(S,'capex',c)}/{RS('Revenue Buildup','nsv',y)}" if (y in FCST or M['capex'].get(y) is not None) else "",PC)
    cset(ws,RM[S]["turn"],c,f"={RS('Revenue Buildup','nsv',y)}/{SAME(S,'ppe_rep',c)}",NUM1)

# ---- Lease Schedule ----
S="Lease Schedule"; ws=WS[S]
for y in HIST:
    cset(ws,RM[S]["close"],YC[y],round(M["lease_liab"][y],1),NUM,bold=True)
    cset(ws,RM[S]["rou"],YC[y],round(M["rou"][y],1),NUM)
cset(ws,RM["Lease Schedule"]["int"],"E",40,NUM); cset(ws,RM["Lease Schedule"]["int"],"D",21,NUM)
cset(ws,RM["Lease Schedule"]["pay"],"E",-137,NUM); cset(ws,RM["Lease Schedule"]["pay"],"D",-126,NUM)
for y in FCST:
    c=YC[y]; pc=prevcol(c)
    cset(ws,RM[S]["open"],c,f"={SAME(S,'close',pc)}",NUM)
    cset(ws,RM[S]["new"],c,f"={RS('PPE Schedule','rou_add',y)}",NUM)
    cset(ws,RM[S]["int"],c,f"={SAME(S,'open',c)}*{RS('Assumptions','lease_int_rate',y)}",NUM)
    cset(ws,RM[S]["pay"],c,f"=-{SAME(S,'open',c)}*{RS('Assumptions','lease_pay_pct',y)}",NUM)
    cset(ws,RM[S]["close"],c,f"={SAME(S,'open',c)}+{SAME(S,'new',c)}+{SAME(S,'int',c)}+{SAME(S,'pay',c)}",NUM,bold=True)
    cset(ws,RM[S]["rou"],c,f"={RS('PPE Schedule','rou_close',y)}",NUM)
    cset(ws,RM[S]["intpl"],c,f"={SAME(S,'int',c)}",NUM)

# ---- Debt & Cash ----
S="Debt & Cash"; ws=WS[S]
for y in HIST:
    cset(ws,RM[S]["borrow"],YC[y],M["borrowings"][y],NUM)
    cset(ws,RM[S]["cash"],YC[y],round(M["cash"][y],1),NUM)
    cset(ws,RM[S]["dep"],YC[y],round(M["bank_deposits"][y],1),NUM)
    cset(ws,RM[S]["inv"],YC[y],round(M["cur_investments"][y],1),NUM)
for y in YEARS:
    c=YC[y]
    cset(ws,RM[S]["lease"],c,f"={RS('Lease Schedule','close',y)}",NUM)
    cset(ws,RM[S]["totdebt"],c,f"={SAME(S,'borrow',c)}+{SAME(S,'lease',c)}",NUM,bold=True)
    cset(ws,RM[S]["treas"],c,f"={RS('Balance Sheet','treas',y)}",NUM,bold=True)
    cset(ws,RM[S]["netdebt"],c,f"={SAME(S,'totdebt',c)}-{SAME(S,'treas',c)}",NUM,bold=True)
for y in FCST:
    c=YC[y]
    cset(ws,RM[S]["borrow"],c,6,NUM)
    cset(ws,RM[S]["cash"],c,f"={SAME(S,'treas',c)}*{RSa('Assumptions','cash_split')}",NUM)
    cset(ws,RM[S]["dep"],c,f"={SAME(S,'treas',c)}*{RSa('Assumptions','dep_split')}",NUM)
    cset(ws,RM[S]["inv"],c,f"={SAME(S,'treas',c)}*{RSa('Assumptions','inv_split')}",NUM)

# ---- Balance Sheet ----
S="Balance Sheet"; ws=WS[S]
hist_inputs={"gw":"goodwill","intang":"intangibles","invprop":"inv_property","jv":"jv_invest",
 "ofa":"other_fin_assets","dta":"dta_net","cta":"cur_tax_assets","oa":"other_assets","hfsa":"assets_hfs",
 "eq":"equity","dtl":"dtl","prov":"provisions","ctl":"cur_tax_liab","ofl":"other_fin_liab",
 "ocl":"other_cur_liab","hfsl":"liab_hfs"}
for y in HIST:
    for k,src in hist_inputs.items():
        cset(ws,RM[S][k],YC[y],round(M[src][y],1),NUM,bold=(k=="eq"))
# links (all years)
for y in YEARS:
    c=YC[y]
    cset(ws,RM[S]["ppe"],c,f"={RS('PPE Schedule','ppe_rep',y)}",NUM)
    cset(ws,RM[S]["cwip"],c,f"={RS('PPE Schedule','cwip',y)}",NUM)
    cset(ws,RM[S]["inv"],c,f"={RS('Working Capital','inv',y)}",NUM)
    cset(ws,RM[S]["recv"],c,f"={RS('Working Capital','recv',y)}",NUM)
    cset(ws,RM[S]["recvnc"],c,f"={RS('Working Capital','recvnc',y)}",NUM)
    cset(ws,RM[S]["borrow"],c,f"={RS('Debt & Cash','borrow',y)}",NUM)
    cset(ws,RM[S]["lease"],c,f"={RS('Lease Schedule','close',y)}",NUM)
    cset(ws,RM[S]["pay"],c,f"={RS('Working Capital','pay',y)}",NUM)
    # non-cash assets subtotal
    nca=("+".join(SAME(S,k,c) for k in ["ppe","cwip","gw","intang","invprop","jv","inv","recv","recvnc","ofa","dta","cta","oa","hfsa"]))
    cset(ws,RM[S]["nca"],c,f"={nca}",NUM)
    tl=("+".join(SAME(S,k,c) for k in ["borrow","lease","pay","dtl","prov","ctl","ofl","ocl","hfsl"]))
    cset(ws,RM[S]["tl"],c,f"={tl}",NUM)
    cset(ws,RM[S]["tle"],c,f"={SAME(S,'eq',c)}+{SAME(S,'tl',c)}",NUM,bold=True)
    cset(ws,RM[S]["treas"],c,f"={SAME(S,'tle',c)}-{SAME(S,'nca',c)}",NUM,bold=True)
    cset(ws,RM[S]["ta"],c,f"={SAME(S,'nca',c)}+{SAME(S,'treas',c)}",NUM,bold=True)
    cset(ws,RM[S]["chk"],c,f"={SAME(S,'ta',c)}-{SAME(S,'tle',c)}",NUM1)
# forecast roll-forwards
for y in FCST:
    c=YC[y]; pc=prevcol(c)
    cset(ws,RM[S]["gw"],c,f"={SAME(S,'gw',pc)}",NUM)
    cset(ws,RM[S]["intang"],c,f"={SAME(S,'intang',pc)}-{RSa('Assumptions','intangible_amort')}",NUM)
    cset(ws,RM[S]["invprop"],c,f"={SAME(S,'invprop',pc)}",NUM)
    cset(ws,RM[S]["jv"],c,f"={SAME(S,'jv',pc)}",NUM)
    cset(ws,RM[S]["ofa"],c,f"={SAME(S,'ofa',pc)}",NUM)
    cset(ws,RM[S]["dta"],c,f"={SAME(S,'dta',pc)}",NUM)
    cset(ws,RM[S]["cta"],c,0,NUM)
    cset(ws,RM[S]["oa"],c,f"={RSa('Assumptions','tax_protest')}+{RSa('Assumptions','other_assets_base')}*{RS('Revenue Buildup','nsv',y)}/{RS('Revenue Buildup','nsv','FY26')}",NUM)
    cset(ws,RM[S]["hfsa"],c,0,NUM)
    cset(ws,RM[S]["div"],c,f"={RS('Income Statement','pat',y)}*{RS('Assumptions','payout',y)}",NUM)
    cset(ws,RM[S]["eq"],c,f"={SAME(S,'eq',pc)}+{RS('Income Statement','pat',y)}-{SAME(S,'div',c)}",NUM,bold=True)
    cset(ws,RM[S]["dtl"],c,f"={SAME(S,'dtl',pc)}",NUM)
    cset(ws,RM[S]["prov"],c,f"={RS('Revenue Buildup','nsv',y)}*{RS('Assumptions','provisions_pct_nsv',y)}",NUM)
    cset(ws,RM[S]["ctl"],c,f"={SAME(S,'ctl',pc)}*{RSa('Assumptions','curtax_growth')}",NUM)
    cset(ws,RM[S]["ofl"],c,f"={SAME(S,'ofl',pc)}",NUM)
    cset(ws,RM[S]["ocl"],c,f"={RS('Revenue Buildup','nsv',y)}*{RS('Assumptions','other_cl_pct_nsv',y)}",NUM)
    cset(ws,RM[S]["hfsl"],c,0,NUM)

# ---- Cash Flow ----
S="Cash Flow"; ws=WS[S]; B="Balance Sheet"
for y in FCST:
    c=YC[y]; pc=prevcol(c)
    cset(ws,RM[S]["pat"],c,f"={RS('Income Statement','pat',y)}",NUM)
    cset(ws,RM[S]["da"],c,f"={RS('Income Statement','dep',y)}",NUM)
    cset(ws,RM[S]["lint"],c,f"={RS('Lease Schedule','int',y)}",NUM)
    dwc=(f"=({RS(B,'pay',y)}-{RS(B,'pay',YEARS[YEARS.index(y)-1])})+({RS(B,'prov',y)}-{RS(B,'prov',YEARS[YEARS.index(y)-1])})"
         f"+({RS(B,'ctl',y)}-{RS(B,'ctl',YEARS[YEARS.index(y)-1])})+({RS(B,'ocl',y)}-{RS(B,'ocl',YEARS[YEARS.index(y)-1])})"
         f"-({RS(B,'inv',y)}-{RS(B,'inv',YEARS[YEARS.index(y)-1])})-({RS(B,'recv',y)}-{RS(B,'recv',YEARS[YEARS.index(y)-1])})"
         f"-({RS(B,'recvnc',y)}-{RS(B,'recvnc',YEARS[YEARS.index(y)-1])})-({RS(B,'oa',y)}-{RS(B,'oa',YEARS[YEARS.index(y)-1])})")
    cset(ws,RM[S]["dwc"],c,dwc,NUM)
    cset(ws,RM[S]["cfo"],c,f"={SAME(S,'pat',c)}+{SAME(S,'da',c)}+{SAME(S,'lint',c)}+{SAME(S,'dwc',c)}",NUM,bold=True)
    cset(ws,RM[S]["capex"],c,f"=-{RS('PPE Schedule','capex',y)}",NUM)
    cset(ws,RM[S]["rcb"],c,f"=({RS(B,'hfsa',YEARS[YEARS.index(y)-1])}-{RS(B,'hfsa',y)})-({RS(B,'hfsl',YEARS[YEARS.index(y)-1])}-{RS(B,'hfsl',y)})",NUM)
    cset(ws,RM[S]["cfi"],c,f"={SAME(S,'capex',c)}+{SAME(S,'rcb',c)}",NUM,bold=True)
    cset(ws,RM[S]["divp"],c,f"=-{RS(B,'div',y)}",NUM)
    cset(ws,RM[S]["leasep"],c,f"={RS('Lease Schedule','pay',y)}",NUM)
    cset(ws,RM[S]["dborrow"],c,f"={RS(B,'borrow',y)}-{RS(B,'borrow',YEARS[YEARS.index(y)-1])}",NUM)
    cset(ws,RM[S]["cff"],c,f"={SAME(S,'divp',c)}+{SAME(S,'leasep',c)}+{SAME(S,'dborrow',c)}",NUM,bold=True)
    cset(ws,RM[S]["net"],c,f"={SAME(S,'cfo',c)}+{SAME(S,'cfi',c)}+{SAME(S,'cff',c)}",NUM,bold=True)
    cset(ws,RM[S]["topen"],c,f"={RS(B,'treas',YEARS[YEARS.index(y)-1])}",NUM)
    cset(ws,RM[S]["tclose"],c,f"={SAME(S,'topen',c)}+{SAME(S,'net',c)}",NUM,bold=True)
    cset(ws,RM[S]["tie"],c,f"={SAME(S,'tclose',c)}-{RS(B,'treas',y)}",NUM1)

wb.save("USL_Financial_Model.xlsx")
print("Core sheets written. RM keys:",list(RM.keys()))

# ===================================================================================
# DCF (live: explicit FCFF + premiumisation fade + bridge + sensitivity)
# ===================================================================================
comps=st["comps"]; hist_val=st["hist_val"]; S_base=st["S_base"]; S_bull=st["S_bull"]; S_bear=st["S_bear"]; CMP=st["CMP"]
ISc={"FY27E":"G","FY28E":"H","FY29E":"I","FY30E":"J"}
ws=wb.create_sheet("DCF"); ws.sheet_view.showGridLines=False
ws.cell(row=1,column=1,value="DCF VALUATION — 2-stage FCFF (explicit FY27-30 + premiumisation fade FY31-36)").font=Font(bold=True,color="FFFFFF",size=12)
for j in range(1,13): ws.cell(row=1,column=j).fill=fillH
expl=["C","D","E","F"]; fade=["G","H","I","J","K","L"]; allc=expl+fade
fy=["FY27E","FY28E","FY29E","FY30E","FY31","FY32","FY33","FY34","FY35","FY36"]
ws.column_dimensions["A"].width=34
for j,cl in enumerate(allc):
    ws.column_dimensions[cl].width=10
    h=ws.cell(row=3,column=column_index_from_string(cl),value=fy[j]); h.font=Font(bold=True,color="FFFFFF"); h.fill=fillN; h.alignment=CTR; h.border=border
ws.cell(row=3,column=1,value="\u20b9 crore").font=Font(bold=True,color="FFFFFF"); ws.cell(row=3,column=1).fill=fillN
# rows
R={}
order=[("growth","NSV growth % (fade inputs)"),("nsv","Net Sales Value"),("mgn","EBITDA margin %"),
 ("ebitda","EBITDA"),("da","Depreciation & amortisation"),("ebit","EBIT"),("tax","Tax on EBIT"),
 ("nopat","NOPAT"),("capex","Capex"),("lease","New leases"),("dwc","Increase in working capital"),
 ("fcff","Free Cash Flow to Firm (FCFF)"),("per","Discount period")]
rr=4
for k,lab in order:
    c=ws.cell(row=rr,column=1,value=lab); c.font=Font(bold=(k in("fcff","ebitda","nopat")),size=10,color=BLACK); c.alignment=LFT; c.border=border
    R[k]=rr; rr+=1
def dc(col,row): return f"{col}{row}"
# fade growth inputs (blue) + margins
fade_g=[0.085,0.080,0.075,0.070,0.065,0.060]
m30=M["bev_ebitda"]["FY30E"]/M["net_revenue"]["FY30E"]
fade_m=[round(m30+(0.225-m30)*i/6,4) for i in range(1,7)]
for j,cl in enumerate(fade):
    cset(ws,R["growth"],cl,fade_g[j],PC)
    cset(ws,R["mgn"],cl,fade_m[j],PC)
# explicit columns: link to live statements
for i,cl in enumerate(expl):
    y=["FY27E","FY28E","FY29E","FY30E"][i]; isc=ISc[y]
    cset(ws,R["nsv"],cl,f"='Revenue Buildup'!{isc}{RM['Revenue Buildup']['nsv']}",NUM)
    cset(ws,R["mgn"],cl,f"='Income Statement'!{isc}{RM['Income Statement']['ebitda']}/'Income Statement'!{isc}{RM['Income Statement']['nsv']}",PC)
    cset(ws,R["ebitda"],cl,f"='Income Statement'!{isc}{RM['Income Statement']['ebitda']}",NUM,bold=True)
    cset(ws,R["da"],cl,f"='Income Statement'!{isc}{RM['Income Statement']['dep']}",NUM)
    cset(ws,R["ebit"],cl,f"={dc(cl,R['ebitda'])}-{dc(cl,R['da'])}",NUM)
    cset(ws,R["tax"],cl,f"={dc(cl,R['ebit'])}*'Assumptions'!{isc}{RM['Assumptions']['tax_rate']}",NUM)
    cset(ws,R["nopat"],cl,f"={dc(cl,R['ebit'])}-{dc(cl,R['tax'])}",NUM,bold=True)
    cset(ws,R["capex"],cl,f"='PPE Schedule'!{isc}{RM['PPE Schedule']['capex']}",NUM)
    cset(ws,R["lease"],cl,f"='PPE Schedule'!{isc}{RM['PPE Schedule']['rou_add']}",NUM)
    # increase in WC from balance sheet
    pc=prevcol(isc); B="Balance Sheet"
    nwc=lambda col:f"('{B}'!{col}{RM[B]['inv']}+'{B}'!{col}{RM[B]['recv']}+'{B}'!{col}{RM[B]['recvnc']}-'{B}'!{col}{RM[B]['pay']}-'{B}'!{col}{RM[B]['ocl']})"
    cset(ws,R["dwc"],cl,f"={nwc(isc)}-{nwc(pc)}",NUM)
    cset(ws,R["fcff"],cl,f"={dc(cl,R['nopat'])}+{dc(cl,R['da'])}-{dc(cl,R['capex'])}-{dc(cl,R['lease'])}-{dc(cl,R['dwc'])}",NUM,bold=True)
# fade columns
for j,cl in enumerate(fade):
    pc = "F" if j==0 else fade[j-1]   # prior col (F=FY30)
    pnsv = f"'Revenue Buildup'!J{RM['Revenue Buildup']['nsv']}" if j==0 else dc(pc,R["nsv"])
    cset(ws,R["nsv"],cl,f"={pnsv}*(1+{dc(cl,R['growth'])})",NUM)
    cset(ws,R["ebitda"],cl,f"={dc(cl,R['nsv'])}*{dc(cl,R['mgn'])}",NUM,bold=True)
    cset(ws,R["da"],cl,f"={dc(cl,R['nsv'])}*('Income Statement'!J{RM['Income Statement']['dep']}/'Revenue Buildup'!J{RM['Revenue Buildup']['nsv']})",NUM)
    cset(ws,R["ebit"],cl,f"={dc(cl,R['ebitda'])}-{dc(cl,R['da'])}",NUM)
    cset(ws,R["tax"],cl,f"={dc(cl,R['ebit'])}*0.255",NUM)
    cset(ws,R["nopat"],cl,f"={dc(cl,R['ebit'])}-{dc(cl,R['tax'])}",NUM,bold=True)
    cset(ws,R["capex"],cl,f"={dc(cl,R['nsv'])}*0.018",NUM)
    cset(ws,R["lease"],cl,f"={dc(cl,R['nsv'])}*0.016",NUM)
    cset(ws,R["dwc"],cl,f"={dc(cl,R['nsv'])}*{dc(cl,R['growth'])}*0.16",NUM)
    cset(ws,R["fcff"],cl,f"={dc(cl,R['nopat'])}+{dc(cl,R['da'])}-{dc(cl,R['capex'])}-{dc(cl,R['lease'])}-{dc(cl,R['dwc'])}",NUM,bold=True)
for j,cl in enumerate(allc): cset(ws,R["per"],cl,j+1,NUM)
fcff_rng=f"$C${R['fcff']}:$L${R['fcff']}"; per_rng=f"$C${R['per']}:$L${R['per']}"; lastf=f"$L${R['fcff']}"
# bridge
br=rr+1
ws.cell(row=br,column=1,value="DCF BRIDGE").font=Font(bold=True,color="FFFFFF")
for j in range(1,4): ws.cell(row=br,column=j).fill=fillH
r_wacc,r_g,r_pvex,r_tv,r_pvtv,r_ev,r_nc,r_rcb,r_eq,r_tp = (br+1,br+2,br+3,br+4,br+5,br+6,br+7,br+8,br+9,br+10)
bridge=[(r_wacc,"WACC",f"='Assumptions'!$B${RM['Assumptions']['wacc']}",PC),
 (r_g,"Terminal growth (g)",f"='Assumptions'!$B${RM['Assumptions']['term_g']}",PC),
 (r_pvex,"PV of explicit + fade FCFF",f"=SUMPRODUCT({fcff_rng},(1+$B${r_wacc})^(-{per_rng}))",NUM),
 (r_tv,"Terminal value (FY36)",f"={lastf}*(1+$B${r_g})/($B${r_wacc}-$B${r_g})",NUM),
 (r_pvtv,"PV of terminal value",f"=$B${r_tv}/(1+$B${r_wacc})^10",NUM),
 (r_ev,"Enterprise Value",f"=$B${r_pvex}+$B${r_pvtv}",NUM),
 (r_nc,"Add: Net cash",f"=-'Debt & Cash'!F{RM['Debt & Cash']['netdebt']}",NUM),
 (r_rcb,"Add: RCB optionality",f"='Assumptions'!$B${RM['Assumptions']['rcb_value']}",NUM),
 (r_eq,"Equity Value",f"=$B${r_ev}+$B${r_nc}+$B${r_rcb}",NUM),
 (r_tp,"DCF value per share (\u20b9)",f"=$B${r_eq}/'Assumptions'!$B${RM['Assumptions']['shares']}",RUP),]
for row,lab,f,fmt in bridge:
    c=ws.cell(row=row,column=1,value=lab); c.font=Font(bold=(lab in("Enterprise Value","Equity Value","DCF value per share (\u20b9)")),size=10,color=BLACK); c.alignment=LFT; c.border=border
    cset(ws,row,"B",f,fmt,bold=(lab=="DCF value per share (\u20b9)"),fill=(fillGo if lab=="DCF value per share (\u20b9)" else None))
ws.column_dimensions["B"].width=16
# sensitivity
sr=br+12
ws.cell(row=sr,column=1,value="SENSITIVITY — DCF value/share (\u20b9): WACC (rows) x terminal g (cols)").font=Font(bold=True,color="FFFFFF")
for j in range(1,8): ws.cell(row=sr,column=j).fill=fillH
gs=[0.050,0.055,0.060,0.065,0.070]; waccs=[0.095,0.100,0.105,0.110,0.115]
ws.cell(row=sr+1,column=1,value="WACC \\ g").font=Font(bold=True,color="FFFFFF"); ws.cell(row=sr+1,column=1).fill=fillN; ws.cell(row=sr+1,column=1).border=border
for j,g in enumerate(gs): cset(ws,sr+1,get_column_letter(2+j),g,PC,bold=True,fill=fillN,color="FFFFFF",align=CTR)
nc_ref=f"-'Debt & Cash'!F{RM['Debt & Cash']['netdebt']}"; rcb_ref=f"'Assumptions'!$B${RM['Assumptions']['rcb_value']}"; sh_ref=f"'Assumptions'!$B${RM['Assumptions']['shares']}"
for i,w in enumerate(waccs):
    row=sr+2+i
    cset(ws,row,"A",w,PC,bold=True,fill=fillN,color="FFFFFF",align=CTR)
    for j,g in enumerate(gs):
        col=get_column_letter(2+j); gcell=f"{get_column_letter(2+j)}${sr+1}"; wcell=f"$A${row}"
        f=(f"=(SUMPRODUCT({fcff_rng},(1+{wcell})^(-{per_rng}))"
           f"+({lastf}*(1+{gcell})/({wcell}-{gcell}))/(1+{wcell})^10"
           f"+{nc_ref}+{rcb_ref})/{sh_ref}")
        fill=fillGo if (abs(w-0.1025)<0.004 and abs(g-0.0625)<0.004) else None
        cset(ws,row,col,f,RUP,fill=fill)
ws.cell(row=sr+8,column=1,value="Live formulas: every figure recalculates from the Assumptions sheet. Green = links to other sheets; black = calculations; blue = inputs.").font=Font(italic=True,size=9,color="555555")

# ===================================================================================
# Comps  (market data blue; multiples black; USL operating figures linked green)
# ===================================================================================
ws=wb.create_sheet("Comps"); ws.sheet_view.showGridLines=False
ws.cell(row=1,column=1,value="TRADING COMPARABLES (market data ~10-Jun-2026)").font=Font(bold=True,color="FFFFFF",size=12)
for j in range(1,11): ws.cell(row=1,column=j).fill=fillH
hh=["Company","MktCap","EV","NetRev FY26","EBITDA FY26","PAT FY26","EV/EBITDA","EV/Sales","P/E","EBITDA mgn"]
for j,t in enumerate(hh): cset(ws,2,get_column_letter(1+j),t,bold=True,fill=fillN,color="FFFFFF",align=CTR)
ws.column_dimensions["A"].width=24
r=3
for _,row in comps.iterrows():
    isUSL=row["Company"].startswith("United")
    cset(ws,r,"A",row["Company"],align=LFT,bold=isUSL,fill=(fillGo if isUSL else None))
    cset(ws,r,"B",row["MktCap"],NUM,fill=(fillGo if isUSL else None))
    cset(ws,r,"C",row["EV"],NUM,fill=(fillGo if isUSL else None))
    if isUSL:
        cset(ws,r,"D",f"='Revenue Buildup'!F{RM['Revenue Buildup']['nsv']}",NUM,fill=fillGo)
        cset(ws,r,"E",f"='Income Statement'!F{RM['Income Statement']['ebitda']}",NUM,fill=fillGo)
        cset(ws,r,"F",f"='Income Statement'!F{RM['Income Statement']['patc']}",NUM,fill=fillGo)
    else:
        cset(ws,r,"D",row["NetRev"],NUM); cset(ws,r,"E",row["EBITDA"],NUM); cset(ws,r,"F",row["PAT"],NUM)
    cset(ws,r,"G",f"=C{r}/E{r}",XF,fill=(fillGo if isUSL else None))
    cset(ws,r,"H",f"=C{r}/D{r}",XF,fill=(fillGo if isUSL else None))
    cset(ws,r,"I",f"=B{r}/F{r}",XF,fill=(fillGo if isUSL else None))
    cset(ws,r,"J",f"=E{r}/D{r}",PC,fill=(fillGo if isUSL else None))
    r+=1
cset(ws,r,"A","Peer median (ex-USL)",bold=True,fill=fillL,align=LFT)
for col in ["B","C","D","E","F"]: ws.cell(row=r,column=column_index_from_string(col)).fill=fillL
cset(ws,r,"G",f"=MEDIAN(G4:G{r-1})",XF,bold=True,fill=fillL)
cset(ws,r,"H",f"=MEDIAN(H4:H{r-1})",XF,bold=True,fill=fillL)
cset(ws,r,"I",f"=MEDIAN(I4:I{r-1})",XF,bold=True,fill=fillL)
cset(ws,r,"J",f"=MEDIAN(J4:J{r-1})",PC,bold=True,fill=fillL)
ws.cell(row=r+2,column=1,value="USL NetRev/EBITDA/PAT linked live to model (green). Peers hard-coded from public data; Tilaknagar PAT normalised; ABDL figures approximate.").font=Font(italic=True,size=9,color="555555")

# ===================================================================================
# Historical Valuation
# ===================================================================================
ws=wb.create_sheet("Historical Valuation"); ws.sheet_view.showGridLines=False
ws.cell(row=1,column=1,value="HISTORICAL VALUATION (FY22-FY26)").font=Font(bold=True,color="FFFFFF",size=12)
for j in range(1,7): ws.cell(row=1,column=j).fill=fillH
cset(ws,2,"A","\u20b9 crore / x",bold=True,fill=fillN,color="FFFFFF",align=LFT)
for j,y in enumerate(HIST): cset(ws,2,get_column_letter(2+j),y,bold=True,fill=fillN,color="FFFFFF",align=CTR)
ws.column_dimensions["A"].width=30
mcap={"FY22":64607,"FY23":55010,"FY24":82500,"FY25":101924,"FY26":91639}
ndc={"FY22":-(341.7+263.7-54.5-5.8-222.1),"FY23":-(1.1+182.2-115.1-768.2-255.8),"FY24":-(25+240-1052-217-599),"FY25":-(0+480-1328-702-873),"FY26":-(6+407-859-1118-1157)}
rows=[("mcap","Market capitalisation"),("pat","PAT (reported)"),("pe","P/E (x)"),
      ("nd","Net cash"),("ev","Enterprise value"),("ebitda","Core EBITDA"),("eve","EV/EBITDA (x)")]
rmap={}; rr=3
for k,lab in rows:
    cset(ws,rr,"A",lab,bold=(k in("pe","eve")),align=LFT); rmap[k]=rr; rr+=1
for j,y in enumerate(HIST):
    cl=get_column_letter(2+j); isc=YC[y]
    cset(ws,rmap["mcap"],cl,mcap[y],NUM)
    cset(ws,rmap["pat"],cl,f"='Income Statement'!{isc}{RM['Income Statement']['pat']}",NUM)
    cset(ws,rmap["pe"],cl,f"={cl}{rmap['mcap']}/{cl}{rmap['pat']}",XF,bold=True)
    cset(ws,rmap["nd"],cl,round(ndc[y],1),NUM)
    cset(ws,rmap["ev"],cl,f"={cl}{rmap['mcap']}-{cl}{rmap['nd']}",NUM)
    cset(ws,rmap["ebitda"],cl,f"='Income Statement'!{isc}{RM['Income Statement']['ebitda']}",NUM)
    cset(ws,rmap["eve"],cl,f"={cl}{rmap['ev']}/{cl}{rmap['ebitda']}",XF,bold=True)
ws.cell(row=rr+1,column=1,value="5-yr avg P/E ~60x; current ~50x (below average). PAT & EBITDA linked live to Income Statement (green).").font=Font(italic=True,size=9,color="555555")

# ===================================================================================
# Scenarios (model-run outputs)
# ===================================================================================
ws=wb.create_sheet("Scenarios"); ws.sheet_view.showGridLines=False
ws.cell(row=1,column=1,value="SCENARIO ANALYSIS (FY30E outcomes & 12-m target)").font=Font(bold=True,color="FFFFFF",size=12)
for j in range(1,5): ws.cell(row=1,column=j).fill=fillH
for j,t in enumerate(["Metric","Bear","Base","Bull"]): cset(ws,2,get_column_letter(1+j),t,bold=True,fill=fillN,color="FFFFFF",align=CTR)
ws.column_dimensions["A"].width=30
srows=[("FY30E Net Sales",S_bear["nsv30"],S_base["nsv30"],S_bull["nsv30"],NUM),
 ("FY30E EBITDA",S_bear["ebitda30"],S_base["ebitda30"],S_bull["ebitda30"],NUM),
 ("FY30E PAT",S_bear["pat30"],S_base["pat30"],S_bull["pat30"],NUM),
 ("FY30E EPS (\u20b9)",S_bear["eps30"],S_base["eps30"],S_bull["eps30"],NUM1),
 ("NSV CAGR FY26-30",S_bear["nsv_cagr"],S_base["nsv_cagr"],S_bull["nsv_cagr"],PC),
 ("EPS CAGR FY26-30",S_bear["eps_cagr"],S_base["eps_cagr"],S_bull["eps_cagr"],PC),
 ("DCF value (\u20b9)",S_bear["tp_dcf"],S_base["tp_dcf"],S_bull["tp_dcf"],RUP),
 ("Relative value (\u20b9)",S_bear["tp_rel"],S_base["tp_rel"],S_bull["tp_rel"],RUP),
 ("Target price (\u20b9)",S_bear["tp"],S_base["tp"],S_bull["tp"],RUP),
 ("Upside/(downside) vs \u20b91,259",S_bear["upside"],S_base["upside"],S_bull["upside"],PC)]
rr=3
for lab,b,ba,bu,fmt in srows:
    bold=lab.startswith("Target")
    cset(ws,rr,"A",lab,bold=bold,align=LFT,fill=(fillGo if bold else None))
    for j,v in enumerate([b,ba,bu]):
        cset(ws,rr,get_column_letter(2+j),round(v,2) if fmt in(NUM1,) else (round(v,4) if fmt==PC else round(v)),fmt,bold=bold,fill=(fillGo if bold else None),color=BLUE)
    rr+=1
ws.cell(row=rr+1,column=1,value="Scenario outputs from separate runs of the model engine (bull/base/bear assumption sets). Base ties to the live DCF & relative valuation.").font=Font(italic=True,size=9,color="555555")

# ===================================================================================
# Sources & References
# ===================================================================================
ws=wb.create_sheet("Sources & References"); ws.sheet_view.showGridLines=False
ws.cell(row=1,column=1,value="SOURCES & REFERENCES — provenance of every hard-coded (blue) figure").font=Font(bold=True,color="FFFFFF",size=12)
for j in range(1,6): ws.cell(row=1,column=j).fill=fillH
ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=95
src=[("COLOUR KEY","Blue = hard-coded input | Black = calculation (same-sheet formula) | Green = link to another sheet"),
 ("FY2022 & FY2023 actuals","USL Annual Report 2022-23 (FY2023.pdf): Consolidated Balance Sheet (p.258-259), Statement of Profit & Loss (p.260-261), Cash Flows (p.263-264). Reported in \u20b9 million -> divided by 10 to \u20b9 crore."),
 ("FY2024 & FY2025 actuals","USL Integrated Annual Report 2024-25 (FY2025.pdf): Consolidated Balance Sheet (p.410-411), Statement of P&L (p.412-413), Cash Flows (p.415-416), PPE note 3.1, ROU/Lease note 3.2, Segment note (p.460-461). \u20b9 crore."),
 ("FY2026 actuals","USL Consolidated Financial Results for year ended 31-Mar-2026 (fy 2026.xlsx), filed 14-May-2026: P&L, Balance Sheet, Cash Flow, Segment. Continuing operations (RCB = discontinued/held-for-sale). \u20b9 crore."),
 ("Volume / NSV-per-case / P&A salience","FY2025 AR Key Performance Indicators (p.22-23) and Board's Report (p.92); FY2023 AR MD&A (p.34,48). Standalone NSV used as proxy for beverage-alcohol NSV (consolidation differences immaterial)."),
 ("Net Sales Value (NSV)","Defined as Revenue from operations minus Excise duty. Excise is a pass-through state levy (~55% of gross). Driver of the model topline."),
 ("PPE gross block / depreciation","FY2025 AR note 3.1 (gross block, additions, disposals, accumulated depreciation, useful lives)."),
 ("Lease (Ind AS 116)","FY2025 AR note 3.2 (ROU roll-forward, lease liability current/non-current, lease cash outflows, incremental borrowing rate)."),
 ("Market data (price, mkt cap, peers)","Public market sources ~10-Jun-2026 (NSE/exchange data, company FY26 results). USL CMP \u20b91,259; mkt cap \u20b991,639 cr. Radico/ABDL/Tilaknagar/Globus FY26 net rev, EBITDA, PAT from public filings/results. Approximate; for relative context only."),
 ("Forecast assumptions","Analyst estimates (see Assumptions sheet, all blue). Justified in the accompanying report (USL_Equity_Research_Report.md)."),
 ("WACC / terminal growth","WACC 10.25% (rf ~6.5%, ERP ~6.0%, beta ~0.65, net-cash); terminal g 6.25% (India nominal long-run consumption). Editable on Assumptions sheet."),
 ("Note on RCB","Sports/RCB classified by the company as discontinued/held-for-sale in FY26; excluded from the core forecast and valued separately as optionality (\u20b92,500 cr, conservative placeholder)."),
 ("Disclaimer","Educational/analytical model. Not investment advice. Assumptions are the analyst's estimates and should be stress-tested against quarterly disclosures."),]
rr=3
for k,v in src:
    cset(ws,rr,"A",k,bold=True,align=LFT,fill=fillL)
    c=ws.cell(row=rr,column=2,value=v); c.font=Font(size=10,color=BLACK); c.alignment=LFT; c.border=border; rr+=1

# ===================================================================================
# Cover
# ===================================================================================
ws=wb.create_sheet("Cover"); ws.sheet_view.showGridLines=False
ws.cell(row=1,column=1,value="UNITED SPIRITS LIMITED (NSE: UNITDSPR) — Integrated Financial Model & Initiation").font=Font(bold=True,color="FFFFFF",size=13)
for j in range(1,7): ws.cell(row=1,column=j).fill=fillH
ws.column_dimensions["A"].width=42; ws.column_dimensions["B"].width=42
DCFr = br+10  # DCF value/share row
cov=[("Rating","NEUTRAL / HOLD",None),
 ("Current market price (10-Jun-2026)",1259,RUP),
 ("12-month target price (base, blended)",1160,RUP),
 ("DCF value/share (live link)",f"='DCF'!$B${DCFr}",RUP),
 ("FY26 Net Sales Value (live link)",f"='Revenue Buildup'!F{RM['Revenue Buildup']['nsv']}",NUM),
 ("FY26 core EBITDA (live link)",f"='Income Statement'!F{RM['Income Statement']['ebitda']}",NUM),
 ("FY26 PAT reported (live link)",f"='Income Statement'!F{RM['Income Statement']['pat']}",NUM),
 ("FY30E Net Sales (live link)",f"='Revenue Buildup'!J{RM['Revenue Buildup']['nsv']}",NUM),
 ("FY30E EBITDA (live link)",f"='Income Statement'!J{RM['Income Statement']['ebitda']}",NUM),
 ("Shares outstanding (cr)",f"='Assumptions'!$B${RM['Assumptions']['shares']}",NUM1),
 ("Market capitalisation (\u20b9 cr)",91639,NUM),
 ("Parent","Diageo plc (~56%)",None),
 ("Reporting basis","Consolidated; topline = Net Sales Value (ex-excise)",None),
 ("Colour key","Blue = input | Black = calculation | Green = cross-sheet link",None),
 ("Note","FY26 = last actual; FY27E-FY30E forecast; RCB = discontinued/held-for-sale (optionality)",None)]
rr=3
for k,v,fmt in cov:
    cset(ws,rr,"A",k,bold=True,align=LFT,fill=fillL)
    if fmt: cset(ws,rr,"B",v,fmt,align=LFT)
    else:
        c=ws.cell(row=rr,column=2,value=v); c.font=Font(size=10,color=BLACK); c.alignment=LFT; c.border=border
    rr+=1
ws.cell(row=rr+1,column=1,value="Educational model — not investment advice. All forecast figures are live Excel formulas linked across sheets.").font=Font(italic=True,size=9,color="555555")

# reorder: Cover, Sources, Assumptions, Revenue..., DCF, Comps, Hist Val, Scenarios
order=["Cover","Sources & References","Assumptions","Revenue Buildup","Income Statement","Working Capital",
 "PPE Schedule","Lease Schedule","Debt & Cash","Balance Sheet","Cash Flow","DCF","Comps","Historical Valuation","Scenarios"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
# enable full recalc on load
try:
    wb.calculation.fullCalcOnLoad=True
except Exception: pass
if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
wb.save("USL_Financial_Model.xlsx")
print("Saved full formula workbook. Sheets:",[s.title for s in wb._sheets])
