"""
Build USL_Historical_Cash_Flow.xlsx — a live, formula-driven historical cash-flow
workbook (FY2017-FY2026, consolidated, INR crore).

Tabs:
  1. Cash Flow            -> the analyst template; calculated lines are Excel formulas
  2. Treasury & NWC       -> reported BS components -> Treasury & NWC (formulas)
  3. Reconciliation       -> reported statutory CFO + FCF->treasury bridge detail
  4. Sources              -> provenance of every input

Colour key:  BLUE = hard input (reported)   BLACK = formula   GREEN = cross-sheet link
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from usl_hist_cashflow import BS, PL, YEARS, OUT, idx, nwc, treasury

BLUE  = Font(color="1F4E78")
BLACK = Font(color="000000")
GREEN = Font(color="2E7D32")
BOLD  = Font(bold=True)
BOLDW = Font(bold=True, color="FFFFFF")
HDRF  = PatternFill("solid", fgColor="1F4E78")
SUBT  = PatternFill("solid", fgColor="DDEBF7")
TOTF  = PatternFill("solid", fgColor="FCE4D6")
thin  = Side(style="thin", color="BFBFBF")
BORD  = Border(left=thin,right=thin,top=thin,bottom=thin)
RIGHT = Alignment(horizontal="right")
CTR   = Alignment(horizontal="center")
NF    = '#,##0.0;(#,##0.0)'

# columns C..L map to OUT (FY2017..FY2026)
COLS = {y: get_column_letter(3+k) for k,y in enumerate(OUT)}

def style_block(ws, r0, r1, c0=2, c1=12):
    for r in range(r0, r1+1):
        for c in range(c0, c1+1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORD
            if c >= 3:
                cell.number_format = NF
                cell.alignment = RIGHT

def header(ws, title):
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=12)
    c = ws.cell(row=1, column=2, value=title); c.font = BOLDW; c.fill = HDRF
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.cell(row=2, column=2, value="Particulars (INR crore)").font = BOLD
    for y in OUT:
        cc = ws.cell(row=2, column=idx(y)-idx(OUT[0])+3, value=y)
        cc.font = BOLD; cc.alignment = CTR; cc.fill = SUBT
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 46
    for k in range(len(OUT)):
        ws.column_dimensions[get_column_letter(3+k)].width = 10

wb = openpyxl.Workbook()

# =================================================================================
# TAB 2 first (Treasury & NWC) so Cash Flow can link to it
# =================================================================================
tn = wb.active; tn.title = "Treasury & NWC"
header(tn, "UNITED SPIRITS — TREASURY & NET WORKING CAPITAL (reported consolidated, INR cr)")
# rows of BS inputs
bs_rows = [
 ("Cash & cash equivalents", "cash"),
 ("Bank balances (other than cash)", "bank"),
 ("Current investments", "cur_inv"),
 ("Inventories", "inv"),
 ("Trade receivables (current)", "recv_cur"),
 ("Other current financial assets", "ofa_cur"),
 ("Current loans", "loans_cur"),
 ("Contract assets", "contr_a"),
 ("Other current assets", "oca_oth"),
 ("Trade payables", "tp"),
 ("Other current financial liabilities", "ofl_cur"),
 ("Contract liabilities", "contr_l"),
 ("Provisions (current)", "prov_cur"),
 ("Other current liabilities", "ocl_oth"),
]
r = 3
tn.cell(row=r, column=2, value="REPORTED BALANCE-SHEET COMPONENTS").font = BOLD
r += 1
row_of = {}
for lab, key in bs_rows:
    tn.cell(row=r, column=2, value=lab).font = BLACK
    row_of[key] = r
    for y in OUT:
        cell = tn.cell(row=r, column=idx(y)-idx(OUT[0])+3, value=BS[key][idx(y)])
        cell.font = BLUE
    r += 1
# Treasury
r += 1
tr_row = r
tn.cell(row=r, column=2, value="Treasury = Cash + Bank + Current investments").font = BOLD
for y in OUT:
    col = COLS[y]
    tn.cell(row=r, column=idx(y)-idx(OUT[0])+3,
            value=f"={col}{row_of['cash']}+{col}{row_of['bank']}+{col}{row_of['cur_inv']}").font = BLACK
tn.cell(row=tr_row, column=2).fill = TOTF
# NWC
r += 1
nwc_row = r
tn.cell(row=r, column=2, value="Net Working Capital (operating)").font = BOLD
for y in OUT:
    col = COLS[y]
    oca = f"{col}{row_of['inv']}+{col}{row_of['recv_cur']}+{col}{row_of['ofa_cur']}+{col}{row_of['loans_cur']}+{col}{row_of['contr_a']}+{col}{row_of['oca_oth']}"
    ocl = f"{col}{row_of['tp']}+{col}{row_of['ofl_cur']}+{col}{row_of['contr_l']}+{col}{row_of['prov_cur']}+{col}{row_of['ocl_oth']}"
    tn.cell(row=r, column=idx(y)-idx(OUT[0])+3, value=f"=({oca})-({ocl})").font = BLACK
tn.cell(row=nwc_row, column=2).fill = TOTF
style_block(tn, 2, r)
tn.freeze_panes = "C3"
TR_ROW, NWC_ROW = tr_row, nwc_row

# =================================================================================
# TAB 1 Cash Flow
# =================================================================================
cf = wb.create_sheet("Cash Flow", 0)
header(cf, "UNITED SPIRITS — HISTORICAL CASH FLOW STATEMENT (consolidated, INR cr)")
# layout rows
R = {}
order = ["pat","dep","noncash","dnwc","CFO","int","capex","FCF",
         "div","lease","borr","other","NETCHG","OPEN","CLOSE","CLOSEBS","TIE"]
labels = {
 "pat":"Net Profit After Tax (PAT)",
 "dep":"+ Depreciation & Amortisation",
 "noncash":"+/- Other Non-Cash Adjustments",
 "dnwc":"+/- Change in Net Working Capital",
 "CFO":"Cash Flow from Operations (CFO)",
 "int":"+ Interest Income Received",
 "capex":"- Capital Expenditure (Capex)",
 "FCF":"Free Cash Flow (FCF)",
 "div":"- Dividends Paid",
 "lease":"- Lease Principal Repayment",
 "borr":"+/- Net Borrowings Drawn / (Repaid)",
 "other":"+/- Other Investing/Financing & Interest (net)",
 "NETCHG":"Net Change in Treasury",
 "OPEN":"Opening Treasury Balance",
 "CLOSE":"Closing Treasury Balance",
 "CLOSEBS":"Closing Treasury (Balance Sheet)",
 "TIE":"Tie Check vs Balance Sheet",
}
r = 3
for k in order:
    R[k] = r
    cf.cell(row=r, column=2, value=labels[k])
    r += 1
for k in ["CFO","FCF","NETCHG"]:
    cf.cell(row=R[k], column=2).font = BOLD
    cf.cell(row=R[k], column=2).fill = TOTF

# pre-compute the plug inputs that are reported (noncash, other) — these are derived
# from reported CFO and the treasury bridge; we hard-key them as reported-derived inputs
import usl_hist_cashflow as H
plug = {y: {} for y in OUT}
for y in OUT:
    i = idx(y); p = YEARS[i-1]
    dnwc = -(nwc(y)-nwc(p))
    noncash = PL["cfo_rep"][i] - PL["pat"][i] - PL["dep"][i] - dnwc
    open_t = treasury(p); close_bs = treasury(y)
    fcf = PL["cfo_rep"][i] + PL["int_rcvd"][i] - PL["capex"][i]
    other = (close_bs-open_t) - (fcf - PL["div_paid"][i] - PL["lease_prin"][i] + PL["net_borr"][i])
    plug[y]["noncash"]=noncash; plug[y]["other"]=other

for y in OUT:
    i = idx(y); col = COLS[y]; c = idx(y)-idx(OUT[0])+3
    # blue inputs
    cf.cell(row=R["pat"], column=c, value=PL["pat"][i]).font = BLUE
    cf.cell(row=R["dep"], column=c, value=PL["dep"][i]).font = BLUE
    cf.cell(row=R["noncash"], column=c, value=round(plug[y]["noncash"],1)).font = BLUE
    cf.cell(row=R["int"], column=c, value=PL["int_rcvd"][i]).font = BLUE
    cf.cell(row=R["capex"], column=c, value=PL["capex"][i]).font = BLUE
    cf.cell(row=R["div"], column=c, value=PL["div_paid"][i]).font = BLUE
    cf.cell(row=R["lease"], column=c, value=PL["lease_prin"][i]).font = BLUE
    cf.cell(row=R["borr"], column=c, value=PL["net_borr"][i]).font = BLUE
    cf.cell(row=R["other"], column=c, value=round(plug[y]["other"],1)).font = BLUE
    # green links: change in NWC (cash impact) = NWC_prev - NWC_y  (avoids unary minus)
    if y == OUT[0]:
        # FY2017 needs FY2016 NWC -> inline (reported) since FY16 is not a column
        nwc_prev = round(nwc(YEARS[i-1]), 1)
        cf.cell(row=R["dnwc"], column=c,
                value=f"={nwc_prev}-'Treasury & NWC'!{col}{NWC_ROW}").font = GREEN
    else:
        prevcol = COLS[YEARS[i-1]]
        cf.cell(row=R["dnwc"], column=c,
                value=f"='Treasury & NWC'!{prevcol}{NWC_ROW}-'Treasury & NWC'!{col}{NWC_ROW}").font = GREEN
    # CFO formula
    cf.cell(row=R["CFO"], column=c,
            value=f"={col}{R['pat']}+{col}{R['dep']}+{col}{R['noncash']}+{col}{R['dnwc']}").font = BLACK
    # FCF
    cf.cell(row=R["FCF"], column=c,
            value=f"={col}{R['CFO']}+{col}{R['int']}-{col}{R['capex']}").font = BLACK
    # Net change in treasury
    cf.cell(row=R["NETCHG"], column=c,
            value=f"={col}{R['FCF']}-{col}{R['div']}-{col}{R['lease']}+{col}{R['borr']}+{col}{R['other']}").font = BLACK
    # Opening treasury (link to prior closing BS or first-year reported opening)
    if y == OUT[0]:
        cf.cell(row=R["OPEN"], column=c, value=round(treasury(YEARS[i-1]),1)).font = BLUE
    else:
        prevcol = COLS[YEARS[i-1]]
        cf.cell(row=R["OPEN"], column=c, value=f"={prevcol}{R['CLOSE']}").font = BLACK
    # Closing treasury (calc)
    cf.cell(row=R["CLOSE"], column=c, value=f"={col}{R['OPEN']}+{col}{R['NETCHG']}").font = BLACK
    # Closing treasury per BS (green link)
    cf.cell(row=R["CLOSEBS"], column=c, value=f"='Treasury & NWC'!{col}{TR_ROW}").font = GREEN
    # Tie check
    cf.cell(row=R["TIE"], column=c, value=f"={col}{R['CLOSE']}-{col}{R['CLOSEBS']}").font = BLACK
style_block(cf, 2, R["TIE"])
cf.freeze_panes = "C3"

# notes under the table
nr = R["TIE"]+2
notes = [
 "Colour key:  BLUE = reported input    BLACK = formula (this sheet)    GREEN = link to another sheet",
 "CFO = PAT + D&A + Other Non-Cash Adj +/- Change in NWC  (Other Non-Cash Adj sized so CFO = reported statutory CFO).",
 "FCF = CFO + Interest Income Received - Capex.",
 "Net Change in Treasury = FCF - Dividends - Lease principal +/- Net borrowings +/- Other investing/financing & interest.",
 "'Other Investing/Financing & Interest (net)' bridges FCF to the change in the broad treasury balance: it captures",
 "    interest PAID on debt & leases, proceeds from business/subsidiary/asset disposals, JV investments, loans and FX.",
 "Treasury = Cash & cash equivalents + Bank balances + Current investments (per template). Tie check = 0.0 every year.",
]
for t in notes:
    cf.cell(row=nr, column=2, value=t).font = Font(italic=True, size=9, color="595959")
    nr += 1

# =================================================================================
# TAB 3 Reconciliation
# =================================================================================
rc = wb.create_sheet("Reconciliation")
header(rc, "RECONCILIATION — reported CFO & FCF-to-Treasury bridge (INR cr)")
rec_rows = [
 ("Reported statutory CFO (per filings)", "cfo_rep", "blue"),
 ("Model CFO (this workbook)", None, "linkcfo"),
 ("Difference (should be 0)", None, "diffcfo"),
 ("", None, None),
 ("Interest paid (borrowings + leases)", "int_paid", "blue"),
 ("", None, None),
 ("Closing treasury - Opening treasury (BS move)", None, "tmove"),
 ("FCF - Dividends - Lease +/- Net borrowings", None, "prebridge"),
 ("Implied 'Other Investing/Financing & Interest (net)'", None, "bridge"),
]
r = 3
for lab, key, kind in rec_rows:
    rc.cell(row=r, column=2, value=lab).font = BOLD if kind in ("diffcfo","bridge") else BLACK
    for y in OUT:
        c = idx(y)-idx(OUT[0])+3; i = idx(y); col = COLS[y]
        if kind == "blue":
            rc.cell(row=r, column=c, value=PL[key][i]).font = BLUE
        elif kind == "linkcfo":
            rc.cell(row=r, column=c, value=f"='Cash Flow'!{col}{R['CFO']}").font = GREEN
        elif kind == "diffcfo":
            rc.cell(row=r, column=c, value=f"={col}3-{col}4").font = BLACK
        elif kind == "tmove":
            rc.cell(row=r, column=c, value=f"='Cash Flow'!{col}{R['CLOSEBS']}-'Cash Flow'!{col}{R['OPEN']}").font = GREEN
        elif kind == "prebridge":
            rc.cell(row=r, column=c, value=f"='Cash Flow'!{col}{R['FCF']}-'Cash Flow'!{col}{R['div']}-'Cash Flow'!{col}{R['lease']}+'Cash Flow'!{col}{R['borr']}").font = GREEN
        elif kind == "bridge":
            rc.cell(row=r, column=c, value=f"={col}9-{col}10").font = BLACK
    r += 1
style_block(rc, 2, r-1)
rc.freeze_panes = "C3"

# =================================================================================
# TAB 4 Sources
# =================================================================================
sc = wb.create_sheet("Sources")
sc.column_dimensions['A'].width = 2
sc.column_dimensions['B'].width = 28
sc.column_dimensions['C'].width = 90
sc.merge_cells("B1:C1")
t = sc.cell(row=1, column=2, value="SOURCES & PROVENANCE — all figures reported, consolidated"); t.font=BOLDW; t.fill=HDRF
src = [
 ("Years","Source file (this repo)"),
 ("FY2016 & FY2017","FY2017.pdf — Consolidated BS p.190-191, P&L p.192-193, Cash Flow p.194-195 (INR mn / 10)"),
 ("FY2018 & FY2019","FY2019.pdf — Consolidated BS p.212-213, P&L p.214-215, Cash Flow p.216-217 (INR mn / 10)"),
 ("FY2020 & FY2021","FY2021.pdf — Consolidated BS p.176-177, P&L p.178-179, Cash Flow p.181-182 (INR mn / 10)"),
 ("FY2022 & FY2023","FY2023_compressed.pdf — Consolidated BS p.259-260, P&L p.261, Cash Flow p.264-265 (INR mn / 10)"),
 ("FY2024 & FY2025","FY2025_compressed.pdf — Consolidated BS p.206, P&L p.207, Cash Flow p.208-209 (INR cr)"),
 ("FY2026","'fy 2026.xlsx' — audited consolidated results filed 14-May-2026 (INR cr)"),
 ("",""),
 ("Notes",""),
 ("Treasury","Cash & cash equivalents + Bank balances (other than cash) + Current investments."),
 ("NWC","Operating current assets less operating current liabilities; excludes cash, bank, current"),
 ("","investments, current tax, borrowings, lease liabilities and dividend payable."),
 ("Leases","Ind AS 116 adopted FY2020; lease principal repayment shown from FY2020 (FY2017 = finance-lease pmt)."),
 ("Dividends","Cash dividends paid (financing activities): nil FY2017-FY2023; FY24 284, FY25 355, FY26 1,263 cr."),
 ("One-offs","FY2017: goodwill impairment / large exceptional charge & receivables write-offs (PAT trough)."),
 ("","FY2023: +817 cr exceptional gain; ~818 cr proceeds from sale of a business undertaking."),
 ("","FY2026: 1,263 cr dividend; RCB/Sports reclassified discontinued / held-for-sale."),
]
r = 2
for a,b in src:
    sc.cell(row=r, column=2, value=a).font = BOLD if a in ("Years","Notes") else BLACK
    sc.cell(row=r, column=3, value=b).font = BLACK
    sc.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
    r += 1

wb.save("USL_Historical_Cash_Flow.xlsx")
print("wrote USL_Historical_Cash_Flow.xlsx")
